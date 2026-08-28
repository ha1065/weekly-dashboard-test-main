#!/usr/bin/env python3
"""
Generate Excel template for forecast data entry.
This creates a pre-populated template with users and projects from the database.
"""

import sys
from pathlib import Path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from src.database.config import SessionLocal
from src.database.models import ClockifyUser, ClockifyProject


def get_monday_of_week(date=None):
    """Get Monday of the week for a given date."""
    if date is None:
        date = datetime.now()
    days_since_monday = date.weekday()
    monday = date - timedelta(days=days_since_monday)
    return monday.replace(hour=0, minute=0, second=0, microsecond=0)


def generate_forecast_template(weeks_forward: int = 12, output_path: str = None):
    """
    Generate an Excel template for forecast data entry.

    Args:
        weeks_forward: Number of weeks to include in the template
        output_path: Path for output file (default: downloads folder)
    """
    db = SessionLocal()

    try:
        # Get active users
        users = db.query(ClockifyUser).filter(
            ClockifyUser.status == 'active',
            ClockifyUser.daily_capacity > 0
        ).order_by(ClockifyUser.name).all()

        # Get active projects
        projects = db.query(ClockifyProject).filter(
            ClockifyProject.archived == False
        ).order_by(ClockifyProject.name).all()

        # Generate week dates
        current_monday = get_monday_of_week()
        weeks = []
        for i in range(weeks_forward):
            week_start = current_monday + timedelta(weeks=i)
            week_end = week_start + timedelta(days=6)
            weeks.append({
                'week_start': week_start.date(),
                'week_end': week_end.date(),
                'label': f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}"
            })

        # Create workbook
        wb = Workbook()

        # === Sheet 1: Forecast Entry ===
        ws_forecast = wb.active
        ws_forecast.title = "Forecast Entry"

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Week Start Date", "Week End Date", "Week Label",
            "User Name", "Clockify User ID", "Location", "Practice Area",
            "Project Name", "Clockify Project ID", "Client Name",
            "Forecasted Hours", "Notes", "Created By"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_forecast.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Add sample rows with validation
        row = 2
        for week in weeks[:4]:  # First 4 weeks as examples
            for user in users[:3]:  # First 3 users as examples
                ws_forecast.cell(row=row, column=1, value=week['week_start'])
                ws_forecast.cell(row=row, column=2, value=week['week_end'])
                ws_forecast.cell(row=row, column=3, value=week['label'])
                ws_forecast.cell(row=row, column=4, value=user.name)
                ws_forecast.cell(row=row, column=5, value=user.clockify_user_id)
                ws_forecast.cell(row=row, column=6, value=user.location or "Unknown")
                ws_forecast.cell(row=row, column=7, value=user.practice_alignment or "")
                ws_forecast.cell(row=row, column=8, value="")  # Project to be filled
                ws_forecast.cell(row=row, column=9, value="")  # Project ID
                ws_forecast.cell(row=row, column=10, value="")  # Client
                ws_forecast.cell(row=row, column=11, value=0)  # Hours
                ws_forecast.cell(row=row, column=12, value="")  # Notes
                ws_forecast.cell(row=row, column=13, value="")  # Created By
                row += 1

        # Column widths
        ws_forecast.column_dimensions['A'].width = 14
        ws_forecast.column_dimensions['B'].width = 14
        ws_forecast.column_dimensions['C'].width = 25
        ws_forecast.column_dimensions['D'].width = 25
        ws_forecast.column_dimensions['E'].width = 26
        ws_forecast.column_dimensions['F'].width = 12
        ws_forecast.column_dimensions['G'].width = 20
        ws_forecast.column_dimensions['H'].width = 35
        ws_forecast.column_dimensions['I'].width = 26
        ws_forecast.column_dimensions['J'].width = 20
        ws_forecast.column_dimensions['K'].width = 15
        ws_forecast.column_dimensions['L'].width = 30
        ws_forecast.column_dimensions['M'].width = 20

        # === Sheet 2: Users Reference ===
        ws_users = wb.create_sheet("Users Reference")

        user_headers = ["User Name", "Clockify User ID", "Email", "Location",
                        "Practice Alignment", "POD Assignment", "Daily Capacity"]

        for col, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, user in enumerate(users, 2):
            ws_users.cell(row=row_idx, column=1, value=user.name)
            ws_users.cell(row=row_idx, column=2, value=user.clockify_user_id)
            ws_users.cell(row=row_idx, column=3, value=user.email)
            ws_users.cell(row=row_idx, column=4, value=user.location or "Unknown")
            ws_users.cell(row=row_idx, column=5, value=user.practice_alignment or "")
            ws_users.cell(row=row_idx, column=6, value=user.pod_assignment or "")
            ws_users.cell(row=row_idx, column=7, value=user.daily_capacity)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_users.column_dimensions[col].width = 25

        # === Sheet 3: Projects Reference ===
        ws_projects = wb.create_sheet("Projects Reference")

        project_headers = ["Project Name", "Clockify Project ID", "Client Name", "Billable"]

        for col, header in enumerate(project_headers, 1):
            cell = ws_projects.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, project in enumerate(projects, 2):
            ws_projects.cell(row=row_idx, column=1, value=project.name)
            ws_projects.cell(row=row_idx, column=2, value=project.clockify_project_id)
            ws_projects.cell(row=row_idx, column=3, value=project.client_name or "")
            ws_projects.cell(row=row_idx, column=4, value="Yes" if project.billable else "No")

        for col in ['A', 'B', 'C', 'D']:
            ws_projects.column_dimensions[col].width = 40

        # === Sheet 4: Weeks Reference ===
        ws_weeks = wb.create_sheet("Weeks Reference")

        week_headers = ["Week Start", "Week End", "Week Label"]

        for col, header in enumerate(week_headers, 1):
            cell = ws_weeks.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, week in enumerate(weeks, 2):
            ws_weeks.cell(row=row_idx, column=1, value=week['week_start'])
            ws_weeks.cell(row=row_idx, column=2, value=week['week_end'])
            ws_weeks.cell(row=row_idx, column=3, value=week['label'])

        for col in ['A', 'B', 'C']:
            ws_weeks.column_dimensions[col].width = 25

        # === Sheet 5: Instructions ===
        ws_instructions = wb.create_sheet("Instructions")

        instructions = [
            ("Resource Forecast Template - Instructions", True),
            ("", False),
            ("How to use this template:", True),
            ("1. Go to the 'Forecast Entry' sheet to enter your forecasts", False),
            ("2. Use the reference sheets to look up User IDs and Project IDs", False),
            ("3. Each row represents one resource's hours on one project for one week", False),
            ("4. Fill in the Forecasted Hours column with the expected hours", False),
            ("5. Save the file and upload it through the Streamlit dashboard", False),
            ("", False),
            ("Required Fields:", True),
            ("- Week Start Date: Must be a Monday", False),
            ("- User Name: Must match a user in the system", False),
            ("- Project Name: The project being worked on", False),
            ("- Forecasted Hours: Number between 0 and 80", False),
            ("", False),
            ("Optional Fields:", True),
            ("- Clockify User ID: Auto-populated if User Name matches", False),
            ("- Client Name: Auto-populated if Project matches", False),
            ("- Notes: Any additional context for the forecast", False),
            ("- Created By: Your name for audit purposes", False),
            ("", False),
            ("Tips:", True),
            ("- You can add as many rows as needed", False),
            ("- Delete the sample data before adding your own", False),
            ("- Use the Weeks Reference sheet to copy week dates", False),
            ("- Use copy/paste from the reference sheets for accuracy", False),
        ]

        for row_idx, (text, is_header) in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_idx, column=1, value=text)
            if is_header:
                cell.font = Font(bold=True, size=12)
            ws_instructions.column_dimensions['A'].width = 80

        # Save workbook
        if output_path is None:
            output_path = f"Forecast_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"

        wb.save(output_path)
        print(f"✅ Template generated: {output_path}")
        print(f"   - {len(users)} active users included")
        print(f"   - {len(projects)} projects included")
        print(f"   - {weeks_forward} weeks of dates")

        return output_path

    finally:
        db.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate forecast Excel template")
    parser.add_argument("--weeks", type=int, default=12, help="Number of weeks forward")
    parser.add_argument("--output", type=str, default=None, help="Output file path")

    args = parser.parse_args()

    generate_forecast_template(weeks_forward=args.weeks, output_path=args.output)

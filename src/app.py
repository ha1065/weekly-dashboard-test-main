"""Main Streamlit application."""

import sys
import os
from pathlib import Path

# Add project root to Python path for imports to work
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Load .env so AWS_PROFILE and other env vars are available to boto3
from dotenv import load_dotenv
load_dotenv(project_root / ".env")

import streamlit as st
import streamlit_authenticator as stauth
import pandas as pd
import uuid
from sqlalchemy import func, text
from datetime import datetime, timedelta
from src.database.config import SessionLocal, engine
from src.database.models import ClockifyUser, ClockifyTimeEntry, ClockifyProject, PSResourceForecast, AppUser
from io import BytesIO

# Apply migrations on startup (only runs once per process)
# Uses schema_migrations table to track applied files — each file runs exactly once
if 'migrations_applied' not in st.session_state:
    from src.shared import apply_pending_migrations
    try:
        print(">>> Starting migration tracking...")
        apply_pending_migrations()
        st.session_state.migrations_applied = True
        print(">>> Migration tracking complete.")
    except Exception as e:
        print(f">>> Migration tracking FAILED: {e}")
        import traceback
        traceback.print_exc()

def get_monday_of_week(date=None):
    """Get Monday of the week for a given date."""
    if date is None:
        date = datetime.now()
    
    # weekday(): Monday=0, Tuesday=1, ..., Sunday=6
    days_since_monday = date.weekday()
    monday = date - timedelta(days=days_since_monday)
    return monday.replace(hour=0, minute=0, second=0, microsecond=0)

def get_sunday_of_week(date=None):
    """Get Sunday of the week for a given date."""
    monday = get_monday_of_week(date)
    sunday = monday + timedelta(days=6)
    return sunday

from src.integrations.forecast_import import import_forecasts_from_template

# Page config
st.set_page_config(
    page_title="Weekly Reporting - Cloudelligent",
    page_icon="📊",
    layout="wide"
)

# =============================================================================
# Authentication Configuration
# =============================================================================
# Get credentials from environment variables
# Set these in ECS task definition or .env file:
#   AUTH_USERNAME - login username
#   AUTH_PASSWORD_HASH - bcrypt hashed password
#   AUTH_NAME - display name
#   AUTH_COOKIE_KEY - random secret key for cookie signing
#   DISABLE_AUTH - set to 'true' to disable authentication (local dev only)

# Check if authentication is disabled (for local development)
DISABLE_AUTH = os.environ.get('DISABLE_AUTH', 'false').lower() == 'true'

if not DISABLE_AUTH:
    def get_auth_config():
        """Build authentication config from the app_users database table.

        On first run (empty table), seeds an initial user from env vars so
        existing deployments keep working without manual intervention.
        """
        import bcrypt as _bcrypt
        db = SessionLocal()
        try:
            users = db.query(AppUser).filter(AppUser.is_active == True).all()

            if not users:
                # Seed initial user from env vars
                username = os.environ.get('AUTH_USERNAME', 'admin')
                password_hash = os.environ.get(
                    'AUTH_PASSWORD_HASH',
                    '$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TtxMQJqhN8/X4.G4HgpqJ3KEajGy'
                )
                display_name = os.environ.get('AUTH_NAME', 'Admin User')
                seed = AppUser(
                    username=username,
                    display_name=display_name,
                    password_hash=password_hash,
                    is_active=True,
                )
                db.add(seed)
                db.commit()
                users = [seed]

            credentials = {
                'usernames': {
                    u.username: {'name': u.display_name, 'password': u.password_hash}
                    for u in users
                }
            }
        finally:
            db.close()

        cookie_key = os.environ.get('AUTH_COOKIE_KEY', 'weekly_reporting_secret_key_change_me')
        return {
            'credentials': credentials,
            'cookie': {
                'name': 'weekly_reporting_auth',
                'key': cookie_key,
                'expiry_days': 7,
            },
        }

    auth_config = get_auth_config()

    authenticator = stauth.Authenticate(
        auth_config['credentials'],
        auth_config['cookie']['name'],
        auth_config['cookie']['key'],
        auth_config['cookie']['expiry_days']
    )

    # Login widget - newer API stores results in session state
    authenticator.login()

    # Get authentication status from session state
    authentication_status = st.session_state.get('authentication_status')
    name = st.session_state.get('name')
    username = st.session_state.get('username')

    if authentication_status == False:
        st.error('Username/password is incorrect')
        st.stop()
    elif authentication_status == None:
        st.warning('Please enter your username and password')
        st.stop()
else:
    # Authentication disabled - set default values
    authentication_status = True
    name = 'Local User'
    username = 'local'

# =============================================================================
# Authenticated Content Below
# =============================================================================

# Title
st.title("📊 Cloudelligent Weekly Reporting")
st.divider()

# Sidebar with logo and logout
logo_path = Path(__file__).parent / "assets" / "logo.png"
if logo_path.exists():
    st.sidebar.image(str(logo_path), width=200)
    st.sidebar.markdown("")  # Add spacing after logo

st.sidebar.title("Navigation")
if not DISABLE_AUTH:
    authenticator.logout('Logout', 'sidebar')
st.sidebar.markdown(f"Logged in as: **{name or 'User'}**")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "Go to",
    ["Governance", "Resource Forecast", "Project Config", "Data Management", "AI Analysis", "Settings"]
)

# Initialize session state for week range if not exists
if 'weeks_back' not in st.session_state:
    st.session_state.weeks_back = 4
if 'weeks_forward' not in st.session_state:
    st.session_state.weeks_forward = 0

# Database session
db = SessionLocal()

if page == "Governance":
    st.header("📋 Governance")

    # Data freshness row
    try:
        with engine.connect() as _fc:
            _cw_time = _fc.execute(text("SELECT MAX(synced_at) FROM clockify_detailed_time_entries")).scalar()
            _jira_time = _fc.execute(text("SELECT MAX(imported_at) FROM jira_issues")).scalar()
            _fc_time = _fc.execute(text("SELECT MAX(created_at) FROM ps_resource_forecast")).scalar()
        _f1, _f2, _f3, _f4 = st.columns(4)
        _f1.metric("🕐 Clockify", _cw_time.strftime('%b %d %H:%M') if _cw_time else "—")
        _f2.metric("🎫 Jira", _jira_time.strftime('%b %d %H:%M') if _jira_time else "—")
        _f3.metric("📈 Forecasts", _fc_time.strftime('%b %d %H:%M') if _fc_time else "—")
        _f4.metric("📊 QuickSight", "See Data Mgmt")
    except Exception:
        pass

    # ── Section 1: Compliance — who hasn't logged time this week (S2-11) ──
    try:
        _last_complete_monday = db.execute(text(
            "SELECT (DATE_TRUNC('week', CURRENT_DATE)::DATE - 7)::DATE"
        )).scalar()
        _missing = db.execute(text("""
            SELECT employee_name, pod_assignment, practice_alignment
            FROM vw_weekly_compliance_report
            WHERE is_compliant = 0
            ORDER BY pod_assignment, employee_name
        """)).fetchall()
        if _missing:
            st.error(f"⚠️ **{len(_missing)} staff have not logged time** for week of {_last_complete_monday.strftime('%b %d')}")
            with st.expander(f"View {len(_missing)} non-compliant staff", expanded=False):
                _mc_df = pd.DataFrame(_missing, columns=["Name", "POD", "Practice"])
                st.dataframe(_mc_df, use_container_width=True, hide_index=True)
        else:
            st.success(f"✅ All staff logged time for week of {_last_complete_monday.strftime('%b %d')}")
    except Exception as _e:
        st.caption(f"Compliance data unavailable: {_e}")

    st.divider()

    # ── Section 2: KPI tiles (PS + MC hours only — QuickSight has the detail) ──
    current_monday = get_monday_of_week()
    current_sunday = get_sunday_of_week()

    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        week_option = st.selectbox(
            "Quick Select",
            ["Last Week", "Current Week", "Last 4 Weeks", "Custom Range"],
            key="week_selector"
        )
    with col3:
        refresh = st.button("🔄 Refresh", type="primary")

    if week_option == "Current Week":
        start_date = current_monday.date()
        end_date = current_sunday.date()
    elif week_option == "Last Week":
        start_date = (current_monday - timedelta(weeks=1)).date()
        end_date = (current_sunday - timedelta(weeks=1)).date()
    elif week_option == "Last 4 Weeks":
        start_date = (current_monday - timedelta(weeks=4)).date()
        end_date = current_sunday.date()
    else:
        col_start, col_end = st.columns(2)
        with col_start:
            start_date = st.date_input("Start (Monday)", value=current_monday.date())
        with col_end:
            end_date = st.date_input("End (Sunday)", value=current_sunday.date())

    # Data freshness
    try:
        if _last_sync:
            st.caption(f"⏱ Time data as of {_last_sync.strftime('%a %b %d, %H:%M')} CT")
    except Exception:
        pass

    st.caption(f"📅 {start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}")

    # KPI queries
    entries = db.query(
        ClockifyTimeEntry.practice_alignment,
        func.sum(ClockifyTimeEntry.duration_hours).label("total_hours"),
        func.count(func.distinct(ClockifyTimeEntry.clockify_user_id)).label("resource_count")
    ).filter(ClockifyTimeEntry.entry_date.between(start_date, end_date)
    ).group_by(ClockifyTimeEntry.practice_alignment).all()

    pod_data_raw = db.query(
        ClockifyTimeEntry.pod_assignment,
        func.sum(ClockifyTimeEntry.duration_hours).label("total_hours"),
        func.count(func.distinct(ClockifyTimeEntry.clockify_user_id)).label("resource_count")
    ).filter(
        ClockifyTimeEntry.entry_date.between(start_date, end_date),
        ClockifyTimeEntry.pod_assignment.isnot(None)
    ).group_by(ClockifyTimeEntry.pod_assignment).all()

    def clean_pod_name(p):
        return p.replace('{','').replace('}','').replace('"','').strip() if p else p

    pod_aggregated = {}
    for p in pod_data_raw:
        n = clean_pod_name(p[0])
        if n in pod_aggregated:
            pod_aggregated[n] = (n, pod_aggregated[n][1]+(p[1] or 0), pod_aggregated[n][2]+(p[2] or 0))
        else:
            pod_aggregated[n] = (n, p[1] or 0, p[2] or 0)
    pod_data = list(pod_aggregated.values())

    managed_cloud_pods = ["Alpha", "Bravo", "A2Z", "SurePoint"]
    ps_data = next((e for e in entries if e[0] == "Professional Services"), None)
    free_agent = next((p for p in pod_data if p[0] == "Free Agent"), None)
    ps_total_hours = (ps_data[1] if ps_data else 0) + (free_agent[1] if free_agent else 0)
    ps_total_resources = (ps_data[2] if ps_data else 0) + (free_agent[2] if free_agent else 0)
    mc_total_hours = sum(p[1] or 0 for p in pod_data if p[0] in managed_cloud_pods)
    mc_resource_query = db.query(func.count(func.distinct(ClockifyTimeEntry.clockify_user_id))).filter(
        ClockifyTimeEntry.entry_date.between(start_date, end_date),
        ClockifyTimeEntry.pod_assignment.in_(managed_cloud_pods)
    ).scalar() or 0

    cols = st.columns(2)
    with cols[0]:
        st.metric("Professional Services", f"{ps_total_hours:.0f} hrs", delta=f"{ps_total_resources} resources")
    with cols[1]:
        st.metric("Managed Cloud", f"{mc_total_hours:.0f} hrs", delta=f"{mc_resource_query} resources")

    st.divider()

    # ── QuickSight links (replaces duplicate visualization sections) ──
    st.subheader("📊 Dashboards")
    st.caption("Detailed analytics are in QuickSight:")
    qs_col1, qs_col2, qs_col3 = st.columns(3)
    with qs_col1:
        st.link_button("COO Operational Analysis →", "https://us-east-1.quicksight.aws.amazon.com/sn/dashboards/coo-operational-dashboard-prod")
    with qs_col2:
        st.link_button("Executive Summary →", "https://us-east-1.quicksight.aws.amazon.com/sn/dashboards/coo-executive-analysis-prod")
    with qs_col3:
        st.link_button("KPI Scorecard →", "https://us-east-1.quicksight.aws.amazon.com/sn/analyses/coo-operational-analysis-prod")

    # Project Time Detail (S4-02)
    with st.expander("🔍 Project Time Detail", expanded=False):
        try:
            _ptd_clients = db.execute(text(
                "SELECT DISTINCT client_name FROM vw_project_time_detail WHERE client_name IS NOT NULL ORDER BY client_name"
            )).fetchall()
            _ptd_client_opts = [r[0] for r in _ptd_clients]
        except Exception:
            _ptd_client_opts = []
        _ptd_col1, _ptd_col2, _ptd_col3 = st.columns(3)
        with _ptd_col1:
            _ptd_client = st.selectbox("Client", ["(all)"] + _ptd_client_opts, key="ptd_client")
        with _ptd_col2:
            _ptd_start = st.date_input("Week From", value=(get_monday_of_week() - timedelta(weeks=4)).date(), key="ptd_start")
        with _ptd_col3:
            _ptd_end = st.date_input("Week To", value=get_monday_of_week().date(), key="ptd_end")
        try:
            _ptd_sql = """
                SELECT entry_date, week_start_date, client_name, project_name,
                       task_name, user_name, billable, duration_hours
                FROM vw_project_time_detail
                WHERE week_start_date BETWEEN :s AND :e
                {client_filter}
                ORDER BY entry_date DESC, client_name, user_name
                LIMIT 500
            """.format(client_filter="AND client_name = :client" if _ptd_client != "(all)" else "")
            _ptd_params = {"s": _ptd_start, "e": _ptd_end}
            if _ptd_client != "(all)":
                _ptd_params["client"] = _ptd_client
            _ptd_rows = db.execute(text(_ptd_sql), _ptd_params).fetchall()
            if _ptd_rows:
                _ptd_df = pd.DataFrame(_ptd_rows, columns=["Date", "Week Start", "Client", "Project", "Task", "Resource", "Billable", "Hours"])
                st.dataframe(_ptd_df, use_container_width=True, hide_index=True, height=350)
                st.caption(f"{len(_ptd_rows)} rows (max 500) | {_ptd_df['Hours'].sum():.1f} total hours")
            else:
                st.info("No data for selected filters.")
        except Exception as _ptd_e:
            st.warning(f"Could not load project time detail: {_ptd_e}")

    # Time entries — collapsed by default (detail view, not COO-primary)
    with st.expander("📋 Time Entry Detail", expanded=False):
        # Add filters - expanded to 5 columns
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            filter_practice = st.multiselect(
                "Practice Alignment",
                options=["Professional Services", "Managed Cloud", "IT Service Delivery", "Service Desk"],
                default=[]
            )

        with col2:
            filter_location = st.multiselect(
                "Location",
                options=["Onshore", "Offshore", "Unknown"],
                default=[]
            )

        with col3:
            filter_pod = st.multiselect(
                "POD Assignment",
                options=[p[0] for p in db.query(ClockifyTimeEntry.pod_assignment).distinct().all() if p[0]],
                default=[]
            )

        with col4:
            filter_skill_area = st.multiselect(
                "Skill Area",
                options=db.query(ClockifyTimeEntry.skill_area).distinct().all(),
                default=[],
                format_func=lambda x: x[0] if x and x[0] else "Unknown"
            )

        with col5:
            limit = st.number_input("Show entries", min_value=10, max_value=500, value=50, step=10)

        # Build query with filters
        query = db.query(ClockifyTimeEntry).filter(
            ClockifyTimeEntry.entry_date.between(start_date, end_date)
        )

        if filter_practice:
            query = query.filter(ClockifyTimeEntry.practice_alignment.in_(filter_practice))

        if filter_location:
            query = query.filter(ClockifyTimeEntry.location.in_(filter_location))

        if filter_pod:
            query = query.filter(ClockifyTimeEntry.pod_assignment.in_(filter_pod))

        if filter_skill_area:
            skill_areas = [sa[0] for sa in filter_skill_area if sa and sa[0]]
            if skill_areas:
                query = query.filter(ClockifyTimeEntry.skill_area.in_(skill_areas))

        recent_entries = query.order_by(ClockifyTimeEntry.start_time.desc()).limit(limit).all()

        if recent_entries:
            df = pd.DataFrame([
                {
                    "Date": e.entry_date,
                    "Resource": e.user_name,
                    "Title": e.cloudelligent_title or "N/A",
                    "POD": e.pod_assignment or "N/A",
                    "Project": e.project_name,
                    "Client": e.client_name or "N/A",
                    "Hours": round(e.duration_hours, 2),
                    "Practice": e.practice_alignment,
                    "Skill Area": e.skill_area or "N/A",
                    "Location": e.location,
                    "Billable": "✓" if e.billable else "✗"
                }
                for e in recent_entries
            ])

            st.dataframe(df, use_container_width=True, height=400)
            st.caption(f"Showing {len(recent_entries)} entries | Total: {df['Hours'].sum():.1f} hours")
        else:
            st.info("No time entries found for selected date range and filters.")

elif page == "Data Management":
    st.header("Data Management")

    # ── Section 1: Refresh Controls ──────────────────────────────────
    st.subheader("Refresh Controls")

    # All QuickSight datasets to refresh
    ALL_QUICKSIGHT_DATASETS = {
        # CloudFormation-managed datasets
        "clockify-time-entries-prod": "Time Entries",
        "clockify-resource-utilization-prod": "Resource Utilization",
        "clockify-weekly-summary-prod": "Weekly Summary",
        "clockify-project-tracking-prod": "Project Tracking",
        "clockify-client-summary-prod": "Client Summary",
        "clockify-practice-performance-prod": "Practice Performance",
        "clockify-monthly-summary-prod": "Monthly Summary",
        "clockify-active-resources-prod": "Active Resources",
        "clockify-pod-performance-prod": "Pod Performance",
        "clockify-skill-area-summary-prod": "Skill Area Summary",
        "clockify-daily-activity-trend-prod": "Daily Activity Trend",
        "clockify-import-activity-prod": "Import Activity",
        # Manually-created datasets
        "clockify-missing-time-submissions-prod": "Missing Time Submissions",
        "3bdc816d-4df6-4db7-b3e6-64e230f28f14": "Forecast Over 40 Hours",
        "42098a5b-a94f-41d5-8300-396f1fec66bf": "Forecast Summary",

        "7833b3c6-cec4-4956-b02a-2316198187cb": "Contractor Weekly Trend",
        "8900f5dc-687e-4d5b-9f91-5efd0cd1daed": "PS Resource Forecasts",
        "9224af38-6c8a-40df-bb83-4dccdf493322": "Time Entries (Manual)",
        "ba87671f-4911-47e1-8185-e381eda339d7": "Employee Utilization",
        "c84d2b1f-de9d-42cd-a389-e425a100c4d4": "Contractor Time Summary",
        "fc56c886-f0d2-4935-8b32-f0862325d3f0": "Forecast vs Actual",
        "data-freshness": "Data Freshness",
        "ps-project-status-view": "PS Project Status (View)",
    }

    # Check for AWS credentials
    aws_available = False
    account_id = None
    try:
        import boto3
        sts = boto3.client('sts')
        account_id = sts.get_caller_identity()['Account']
        aws_available = True
    except Exception:
        pass

    if not aws_available:
        st.warning("AWS credentials not configured. Database view refresh and QuickSight refresh require AWS access.")
    else:
        col1, col2, col3 = st.columns(3)
        with col1:
            btn_views = st.button("Refresh Database Views", type="secondary", use_container_width=True,
                                  help="Invoke Lambda to recreate all database views")
        with col2:
            btn_spice = st.button("Refresh QuickSight Datasets", type="secondary", use_container_width=True,
                                  help="Trigger SPICE ingestion for all QuickSight datasets")
        with col3:
            btn_all = st.button("Refresh All", type="primary", use_container_width=True,
                                help="Refresh database views first, then all QuickSight datasets")

        run_views = btn_views or btn_all
        run_spice = btn_spice or btn_all

        # --- Refresh Database Views ---
        if run_views:
            st.markdown("#### Database Views")
            with st.spinner("Invoking Lambda to refresh database views..."):
                try:
                    import boto3
                    import json as _json

                    lambda_client = boto3.client('lambda', region_name='us-east-1')
                    response = lambda_client.invoke(
                        FunctionName='production-clockify-import',
                        InvocationType='RequestResponse',
                        Payload=_json.dumps({"mode": "apply_views"})
                    )
                    payload = _json.loads(response['Payload'].read().decode('utf-8'))

                    if response.get('StatusCode') == 200 and 'FunctionError' not in response:
                        body = _json.loads(payload.get('body', '{}')) if isinstance(payload.get('body'), str) else payload
                        st.success(f"Database views refreshed successfully.")
                        with st.expander("View Lambda Response"):
                            st.code(_json.dumps(body, indent=2))
                    else:
                        st.error("Lambda invocation returned an error.")
                        with st.expander("View Error Details"):
                            st.code(_json.dumps(payload, indent=2))
                except Exception as e:
                    st.error(f"Lambda invocation error: {str(e)}")

        # --- Refresh QuickSight SPICE Datasets ---
        if run_spice:
            st.markdown("#### QuickSight SPICE Datasets")
            try:
                import boto3
                import time as _time

                quicksight = boto3.client('quicksight', region_name='us-east-1')
                total = len(ALL_QUICKSIGHT_DATASETS)

                progress_bar = st.progress(0)
                status_text = st.empty()
                table_placeholder = st.empty()

                # Phase 1: Trigger all ingestions
                status_text.text("Triggering SPICE ingestions...")
                ingestions = {}
                results = {}

                for dataset_id, display_name in ALL_QUICKSIGHT_DATASETS.items():
                    ingestion_id = f"admin-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{dataset_id[:8]}"
                    try:
                        quicksight.create_ingestion(
                            DataSetId=dataset_id,
                            IngestionId=ingestion_id,
                            AwsAccountId=account_id
                        )
                        ingestions[dataset_id] = ingestion_id
                        results[dataset_id] = {"name": display_name, "status": "Running", "detail": ""}
                    except Exception as e:
                        error_msg = str(e)
                        if 'ResourceExistsException' in error_msg or 'already in progress' in error_msg.lower():
                            results[dataset_id] = {"name": display_name, "status": "Running", "detail": "Already in progress"}
                        else:
                            results[dataset_id] = {"name": display_name, "status": "Failed", "detail": error_msg[:80]}

                def build_status_df(results):
                    rows = []
                    for info in results.values():
                        rows.append({"Dataset": info["name"], "Status": info["status"], "Detail": info["detail"]})
                    return pd.DataFrame(rows)

                def _color_spice_status(val):
                    if val in ('Done', 'COMPLETED'):
                        return 'color: #33A94F; font-weight: 600'
                    elif val in ('Failed', 'FAILED', 'CANCELLED'):
                        return 'color: #D74018; font-weight: 600'
                    elif val in ('Running', 'RUNNING', 'INITIALIZED', 'QUEUED'):
                        return 'color: #FF9B00; font-weight: 600'
                    return ''

                triggered_count = len(ingestions)
                status_text.text(f"Triggered {triggered_count}/{total} datasets. Polling for completion...")
                table_placeholder.dataframe(build_status_df(results).style.map(_color_spice_status, subset=['Status']), use_container_width=True, hide_index=True, height=400)

                # Phase 2: Poll for completion
                timeout = 300
                poll_interval = 5
                start = _time.time()
                pending = {did: iid for did, iid in ingestions.items()}

                while pending and (_time.time() - start) < timeout:
                    _time.sleep(poll_interval)
                    newly_done = []
                    for dataset_id, ingestion_id in pending.items():
                        try:
                            resp = quicksight.describe_ingestion(
                                DataSetId=dataset_id,
                                IngestionId=ingestion_id,
                                AwsAccountId=account_id
                            )
                            ing_status = resp['Ingestion']['IngestionStatus']
                            if ing_status == 'COMPLETED':
                                row_count = resp['Ingestion'].get('RowInfo', {}).get('RowsIngested', '')
                                results[dataset_id]["status"] = "Done"
                                results[dataset_id]["detail"] = f"{row_count} rows" if row_count != '' else "Done"
                                newly_done.append(dataset_id)
                            elif ing_status in ('FAILED', 'CANCELLED'):
                                err = resp['Ingestion'].get('ErrorInfo', {}).get('Message', ing_status)
                                results[dataset_id]["status"] = "Failed"
                                results[dataset_id]["detail"] = err[:80]
                                newly_done.append(dataset_id)
                        except Exception as e:
                            results[dataset_id]["status"] = "Failed"
                            results[dataset_id]["detail"] = str(e)[:80]
                            newly_done.append(dataset_id)

                    for did in newly_done:
                        pending.pop(did, None)

                    completed_count = sum(1 for r in results.values() if r["status"] != "Running")
                    progress_bar.progress(completed_count / total)
                    status_text.text(f"Completed {completed_count}/{total} datasets...")
                    table_placeholder.dataframe(build_status_df(results).style.map(_color_spice_status, subset=['Status']), use_container_width=True, hide_index=True, height=400)

                # Final summary
                progress_bar.progress(1.0)
                done_count = sum(1 for r in results.values() if r["status"] == "Done")
                fail_count = sum(1 for r in results.values() if r["status"] == "Failed")
                still_running = sum(1 for r in results.values() if r["status"] == "Running")

                if fail_count == 0 and still_running == 0:
                    status_text.text(f"All {done_count} datasets refreshed successfully.")
                    st.success(f"All {done_count} datasets refreshed successfully.")
                elif still_running > 0:
                    status_text.text(f"{done_count} done, {still_running} still running (timed out), {fail_count} failed.")
                    st.warning(f"{done_count} completed, {still_running} still running (timed out), {fail_count} failed.")
                else:
                    status_text.text(f"{done_count} done, {fail_count} failed.")
                    st.warning(f"{done_count} completed, {fail_count} failed.")

            except Exception as e:
                st.error(f"QuickSight refresh error: {str(e)}")

    st.divider()

    # ── Section 2: Data Sources ──────────────────────────────────────
    st.subheader("Data Sources")

    DATA_SOURCES = [
        ("clockify_users", "Clockify team members and attributes", "users"),
        ("clockify_projects", "Clockify project definitions", "projects"),
        ("clockify_detailed_time_entries", "Individual time entries from Clockify", "time_entries"),
        ("user_skills", "User skill and certification records", None),
        ("ps_resource_forecasts", "Weekly resource forecast allocations", "forecasts"),
        ("ps_project_status", "Jira project status (PS and MC)", "ps_project_status"),
        ("ps_project_mapping", "Jira-to-Clockify project mapping", "ps_project_mapping"),
        ("jira_projects", "Jira project metadata", "jira_projects"),
        ("import_logs", "Data import audit trail", None),
        ("ai_analysis_prompts", "AI analysis prompt templates by category", None),
        ("ai_analysis_by_user", "AI weekly Jira vs Clockify analysis per user", "analyze_project_health"),
        ("ai_analysis_by_project", "AI weekly Jira vs Clockify analysis per project", "analyze_project_health"),
        ("mc_v2_audit_by_customer", "MC V2 Audit progress scorecard per customer", "mc_v2_audit"),
        ("mc_v2_audit_by_phase", "MC V2 Audit progress scorecard per phase", "mc_v2_audit"),
    ]

    try:
        # Get last-updated timestamps from import_logs
        freshness = db.execute(text(
            "SELECT import_category, MAX(completed_at) AS last_import_at "
            "FROM import_logs WHERE status IN ('success', 'partial') "
            "GROUP BY import_category"
        )).fetchall()
        freshness_lookup = {r[0]: r[1] for r in freshness}
    except Exception:
        freshness_lookup = {}

    try:
        source_rows = []
        for tbl_name, description, freshness_key in DATA_SOURCES:
            try:
                count = db.execute(text(f"SELECT COUNT(*) FROM {tbl_name}")).scalar()
            except Exception:
                count = "N/A"

            last_updated = ""
            if freshness_key and freshness_key in freshness_lookup:
                ts = freshness_lookup[freshness_key]
                if ts:
                    last_updated = ts.strftime('%Y-%m-%d %H:%M')

            source_rows.append({
                "Source": tbl_name,
                "Description": description,
                "Records": count,
                "Last Updated": last_updated,
            })

        st.dataframe(pd.DataFrame(source_rows), use_container_width=True, hide_index=True, height=350)

        try:
            view_count = db.execute(text(
                "SELECT COUNT(*) FROM information_schema.views "
                "WHERE table_schema = 'public'"
            )).scalar()
        except Exception:
            view_count = "?"
        st.caption(
            f"{view_count} reporting views and {len(ALL_QUICKSIGHT_DATASETS)} QuickSight datasets configured. "
            "Use Refresh Controls above to update views and SPICE datasets."
        )
    except Exception as e:
        st.error(f"Error loading data sources: {e}")

    st.divider()

elif page == "Project Config":
    st.header("Project Config")

    # ── Project Mapping ───────────────────────────────────────────
    with st.expander("🗺️ Project Mapping", expanded=False):
        st.markdown("Map Jira projects to Clockify clients/projects so actual hours appear on the Project Status dashboard.")

    try:
        db.close()
    except Exception:
        pass
    db = SessionLocal()

    # ── Sync button ───────────────────────────────────────────────
    if st.button("🔄 Sync Clockify Projects & Clients", help="Pull latest projects and clients from Clockify"):
        with st.spinner("Syncing from Clockify..."):
            try:
                from src.integrations.clockify_client import ClockifyClient
                from src.integrations.import_clockify_data import import_projects
                _client = ClockifyClient()
                import_projects(db, _client)
                st.success("Clockify projects synced.")
                st.rerun()
            except Exception as _e:
                st.error(f"Sync failed: {_e}")

    from src.database.models import PSProjectMapping

    # ── Helper: save/update/delete a mapping ──────────────────────
    def _save_mapping(db_session, ps_client, ps_project, clockify_client, selected_projects, existing_mappings, category='PS'):
        def _trigger_project_status_refresh():
            try:
                import boto3, json as _json
                _lambda = boto3.client('lambda', region_name='us-east-1')
                _lambda.invoke(
                    FunctionName='production-clockify-import',
                    InvocationType='Event',
                    Payload=_json.dumps({
                        "mode": "refresh_quicksight_only",
                        "quicksight_dataset_ids": ["ps-project-status-view", "data-freshness"]
                    })
                )
            except Exception:
                pass  # don't block UI if refresh fails

        try:
            # Re-query fresh from DB to avoid stale session objects
            ps_proj_val = ps_project if ps_project else None
            to_delete = db_session.query(PSProjectMapping).filter(
                PSProjectMapping.ps_client_name == ps_client,
                PSProjectMapping.ps_project_name == ps_proj_val,
            ).all()
            for em in to_delete:
                db_session.delete(em)
            db_session.flush()

            if clockify_client == "-- Not Mapped --":
                db_session.commit()
                if existing_mappings:
                    st.success(f"Removed mapping for {ps_client}")
                    _trigger_project_status_refresh()
                    st.rerun()
                return

            # Insert one row per selected project; no selection = client-level (all projects)
            if not selected_projects:
                db_session.add(PSProjectMapping(
                    ps_client_name=ps_client,
                    ps_project_name=ps_project or None,
                    clockify_client_name=clockify_client,
                    clockify_project_name=None,
                    category=category,
                    is_active=True
                ))
            else:
                for proj_name in selected_projects:
                    db_session.add(PSProjectMapping(
                        ps_client_name=ps_client,
                        ps_project_name=ps_project or None,
                        clockify_client_name=clockify_client,
                        clockify_project_name=proj_name,
                        category=category,
                        is_active=True
                    ))

            db_session.commit()
            proj_label = ", ".join(selected_projects) if selected_projects else "(all projects)"
            st.success(f"Saved: {ps_client} -> {clockify_client} / {proj_label}")
            _trigger_project_status_refresh()
            st.rerun()
        except Exception as e:
            db_session.rollback()
            st.error(f"Failed to save: {e}")

    # ── Helper: render mapping rows for a list of projects ────────
    def _render_mapping_tab(projects, mapping_lookup, clockify_client_list, clockify_projects_list, prefix, category='PS'):
        if not projects:
            st.info("No projects found.")
            return

        # Column headers
        hdr = st.columns([3, 3, 3, 1])
        hdr[0].markdown("**Jira Project**")
        hdr[1].markdown("**Clockify Client**")
        hdr[2].markdown("**Clockify Projects**")
        hdr[3].markdown("**Action**")
        st.divider()

        client_options = ["-- Not Mapped --"] + clockify_client_list

        for i, row in enumerate(projects):
            ps_client = row[0]
            ps_project = row[1] or ""
            key = (ps_client.lower(), ps_project.lower())
            current_list = mapping_lookup.get(key, [])

            cols = st.columns([3, 3, 3, 1])

            with cols[0]:
                st.markdown(f"**{ps_client}**")
                if ps_project:
                    st.caption(ps_project)

            with cols[1]:
                current_client = current_list[0].clockify_client_name if current_list else None
                current_client_idx = 0
                if current_client:
                    try:
                        current_client_idx = client_options.index(current_client)
                    except ValueError:
                        pass
                selected_client = st.selectbox(
                    "Clockify Client",
                    options=client_options,
                    index=current_client_idx,
                    key=f"{prefix}_client_{i}",
                    label_visibility="collapsed"
                )

            with cols[2]:
                project_options = [
                    p[1] for p in clockify_projects_list
                    if p[0] == selected_client and p[1]
                ] if selected_client != "-- Not Mapped --" else []
                current_proj_names = [
                    m.clockify_project_name for m in current_list
                    if m.clockify_project_name
                ]
                valid_defaults = [p for p in current_proj_names if p in project_options]
                selected_projects = st.multiselect(
                    "Clockify Projects",
                    options=project_options,
                    default=valid_defaults,
                    key=f"{prefix}_proj_{i}",
                    label_visibility="collapsed",
                    placeholder="(all projects)"
                )

            with cols[3]:
                if st.button("Save", key=f"{prefix}_save_{i}", type="primary"):
                    _save_mapping(db, ps_client, ps_project, selected_client, selected_projects, current_list, category)

    # ── Load data ─────────────────────────────────────────────────
    try:
        # PS projects
        ps_projects = db.execute(text(
            "SELECT DISTINCT client_name, project_name "
            "FROM ps_project_status "
            "WHERE client_name IS NOT NULL "
            "  AND category = 'PS' "
            "ORDER BY client_name, project_name"
        )).fetchall()

        # MC projects
        mc_projects = db.execute(text(
            "SELECT DISTINCT client_name, project_name "
            "FROM ps_project_status "
            "WHERE client_name IS NOT NULL "
            "  AND category = 'MC' "
            "ORDER BY client_name, project_name"
        )).fetchall()

        # Existing mappings
        mappings = db.query(PSProjectMapping).filter(
            PSProjectMapping.is_active == True
        ).order_by(PSProjectMapping.ps_client_name).all()

        # Distinct Clockify clients
        clockify_clients = db.execute(text(
            "SELECT DISTINCT client_name FROM clockify_detailed_time_entries "
            "WHERE client_name IS NOT NULL ORDER BY client_name"
        )).fetchall()
        clockify_client_list = [r[0] for r in clockify_clients]

        # Distinct Clockify client/project pairs
        clockify_projects = db.execute(text(
            "SELECT DISTINCT client_name, project_name FROM clockify_detailed_time_entries "
            "WHERE client_name IS NOT NULL ORDER BY client_name, project_name"
        )).fetchall()
        clockify_projects_list = [(r[0], r[1]) for r in clockify_projects]

        data_loaded = True
    except Exception as e:
        db.rollback()
        st.error(f"Could not load data: {e}")
        data_loaded = False

    if data_loaded:
        # Build mapping lookup — lists so multi-project mappings are preserved
        mapping_lookup = {}
        for m in mappings:
            key = (m.ps_client_name.lower(), (m.ps_project_name or "").lower())
            mapping_lookup.setdefault(key, []).append(m)

        # Count mapped projects
        ps_mapped = sum(1 for r in ps_projects if (r[0].lower(), (r[1] or "").lower()) in mapping_lookup)
        mc_mapped = sum(1 for r in mc_projects if (r[0].lower(), (r[1] or "").lower()) in mapping_lookup)

        # ── Metrics ───────────────────────────────────────────────
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("PS Projects", len(ps_projects))
        with col2:
            st.metric("PS Mapped", ps_mapped)
        with col3:
            st.metric("MC Projects", len(mc_projects))
        with col4:
            st.metric("MC Mapped", mc_mapped)

        st.divider()

        # ── Tabs ──────────────────────────────────────────────────
        tab_ps, tab_mc = st.tabs(["Professional Services", "Managed Services"])

        with tab_ps:
            _render_mapping_tab(ps_projects, mapping_lookup, clockify_client_list, clockify_projects_list, "ps", "PS")

        with tab_mc:
            _render_mapping_tab(mc_projects, mapping_lookup, clockify_client_list, clockify_projects_list, "mc", "MC")

    st.divider()

    # ── Reporting Exclusions ──────────────────────────────────────
    _excl_users2 = db.execute(text(
        "SELECT clockify_user_id, name, practice_area, pod_assignment, "
        "COALESCE(reporting_excluded, FALSE) AS reporting_excluded "
        "FROM clockify_users WHERE status = 'active' ORDER BY name"
    )).fetchall()
    _excl_count2 = sum(1 for u in _excl_users2 if u[4])

    _excl_label = f"🚫 Reporting Exclusions ({_excl_count2} excluded)"
    with st.expander(_excl_label, expanded=False):
        st.caption('Users excluded here are removed from utilization metrics, compliance reports, and KPI calculations.')
        if _excl_count2 > 0:
            st.warning(f"⚠️ {_excl_count2} user(s) excluded from KPI calculations. Verify before Monday import.")

        with st.form('reporting_exclusions_form_pc'):
            col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
            col1.markdown("**Name**"); col2.markdown("**Practice Area**")
            col3.markdown("**POD**"); col4.markdown("**Excluded**")
            st.divider()
            excl_states2 = {}
            for uid, name_, pa, pod, excluded in _excl_users2:
                cols = st.columns([3, 2, 2, 1])
                cols[0].text(name_); cols[1].text(pa or "—"); cols[2].text(pod or "—")
                excl_states2[uid] = cols[3].checkbox("", value=excluded, key=f"excl_pc_{uid}", label_visibility="collapsed")
            if st.form_submit_button("Save", type="primary"):
                updated = 0
                for uid, name_, pa, pod, excluded in _excl_users2:
                    if excl_states2[uid] != excluded:
                        db.execute(text("UPDATE clockify_users SET reporting_excluded = :val WHERE clockify_user_id = :uid"),
                                   {"val": excl_states2[uid], "uid": uid})
                        updated += 1
                db.commit()
                st.success(f"Updated {updated} users.")

    st.divider()

    # ── Compliance Report Recipients (S2-08: moved from Settings) ──
    _recip_count = db.execute(text("SELECT COUNT(*) FROM compliance_report_recipients WHERE is_active=TRUE")).scalar() or 0
    with st.expander(f"📧 Compliance Report Recipients ({_recip_count} active)", expanded=False):
        recipients2 = db.execute(text(
            "SELECT id, email, display_name, report_run, is_active FROM compliance_report_recipients ORDER BY email"
        )).fetchall()

        if recipients2:
            rec_cols = st.columns([2, 2, 1.5, 0.8, 0.8])
            rec_cols[0].markdown("**Email**"); rec_cols[1].markdown("**Display Name**")
            rec_cols[2].markdown("**Run**"); rec_cols[3].markdown("**Active**"); rec_cols[4].markdown("**Action**")
            st.divider()
            for rec_id, email, display_name, report_run, is_active in recipients2:
                cols = st.columns([2, 2, 1.5, 0.8, 0.8])
                cols[0].text(email); cols[1].text(display_name or ""); cols[2].text(report_run or "")
                with cols[3]:
                    new_active = st.toggle("", value=is_active, key=f"toggle_pc_{rec_id}")
                    if new_active != is_active:
                        db.execute(text("UPDATE compliance_report_recipients SET is_active=:v WHERE id=:id"),
                                  {"v": new_active, "id": rec_id})
                        db.commit()
                        st.rerun()
                with cols[4]:
                    if st.button("✕", key=f"del_pc_{rec_id}"):
                        db.execute(text("DELETE FROM compliance_report_recipients WHERE id=:id"), {"id": rec_id})
                        db.commit()
                        st.rerun()

        st.divider()
        st.caption("**Add Recipient**")
        with st.form("add_compliance_recipient_form"):
            email = st.text_input("Email *", placeholder="user@example.com")
            display_name = st.text_input("Display Name (optional)")
            report_run = st.selectbox("Run", ["morning", "noon", "afternoon", "all"])
            if st.form_submit_button("Add"):
                if email and "@" in email:
                    db.execute(text(
                        "INSERT INTO compliance_report_recipients (email, display_name, report_run) "
                        "VALUES (:e, :n, :r) ON CONFLICT (email) DO UPDATE SET report_run=:r, is_active=TRUE"
                    ), {"e": email, "n": display_name or None, "r": report_run})
                    db.commit()
                    st.rerun()
                else:
                    st.error("Valid email required")
        st.caption("⚠️ SES domain verification required before emails will send.")

    # Customer Status Assignments (S4-03)
    with st.expander("👥 Customer Status Assignments", expanded=False):
        try:
            _csa_rows = db.execute(text(
                "SELECT issue_key, client_name, project_name, category, status, priority, "
                "assignment_role, resource_name FROM vw_customer_status_assignments ORDER BY client_name, project_name"
            )).fetchall()
            if _csa_rows:
                _csa_df = pd.DataFrame(_csa_rows, columns=["Issue", "Client", "Project", "Category", "Status", "Priority", "Role", "Resource"])
                st.dataframe(_csa_df, use_container_width=True, hide_index=True)
            else:
                st.info("No customer status assignment data available.")
        except Exception as _csa_e:
            st.warning(f"Could not load customer status assignments: {_csa_e}")

elif page == "AI Analysis":
    st.header("🤖 AI Analysis")

    aws_available = False
    account_id = None
    try:
        import boto3
        sts = boto3.client('sts')
        account_id = sts.get_caller_identity()['Account']
        aws_available = True
    except Exception:
        pass

    if not aws_available:
        st.warning("AWS credentials not configured. AI Analysis requires AWS access.")

    st.subheader("AI Project Health Analysis")
    st.caption(
        "Configure the prompts used by the Monday AI analysis that compares "
        "Jira activity estimates against Clockify actuals. "
        "Changes take effect on the next run."
    )

    from src.database.models import AIAnalysisPrompt

    # Fixed categories with friendly names and dedicated renderers
    _FIXED_TABS = {
        'PS':       'Professional Services Prompts',
        'MC':       'Managed Cloud Prompts',
        'MC_V2':    'MC V2 Audit Prompt',
        'FORECAST': 'Forecast Analysis Prompt',
    }

    # Discover any additional categories in the DB not covered by the fixed tabs
    _all_db_categories = [
        r[0] for r in
        db.query(AIAnalysisPrompt.category).distinct().order_by(AIAnalysisPrompt.category).all()
    ]
    _extra_categories = [c for c in _all_db_categories if c not in _FIXED_TABS]

    _tab_labels = list(_FIXED_TABS.values()) + _extra_categories
    _all_tabs = st.tabs(_tab_labels)
    _fixed_tab_objs  = dict(zip(_FIXED_TABS.keys(), _all_tabs[:len(_FIXED_TABS)]))
    _extra_tab_objs  = dict(zip(_extra_categories, _all_tabs[len(_FIXED_TABS):]))

    ai_ps_tab       = _fixed_tab_objs['PS']
    ai_mc_tab       = _fixed_tab_objs['MC']
    ai_v2_tab       = _fixed_tab_objs['MC_V2']
    ai_forecast_tab = _fixed_tab_objs['FORECAST']

    def _render_prompt_editor(category: str):
        prompts = (
            db.query(AIAnalysisPrompt)
            .filter(AIAnalysisPrompt.category == category, AIAnalysisPrompt.is_active == True)
            .order_by(AIAnalysisPrompt.sequence_order)
            .all()
        )

        # Pad to 4 slots
        texts = {p.sequence_order: p.prompt_text for p in prompts}
        edited = {}
        for seq in range(1, 5):
            edited[seq] = st.text_area(
                f"Prompt {seq}",
                value=texts.get(seq, ""),
                height=120,
                key=f"ai_prompt_{category}_{seq}",
            )

        if st.button(f"Save {category} Prompts", key=f"save_ai_{category}", type="primary"):
            try:
                # Delete all existing for this category then re-insert
                db.query(AIAnalysisPrompt).filter(
                    AIAnalysisPrompt.category == category
                ).delete()
                for seq, text_val in edited.items():
                    if text_val.strip():
                        db.add(AIAnalysisPrompt(
                            category=category,
                            sequence_order=seq,
                            prompt_text=text_val.strip(),
                            is_active=True,
                        ))
                db.commit()
                st.success(f"{category} prompts saved.")
                st.rerun()
            except Exception as exc:
                db.rollback()
                st.error(f"Failed to save prompts: {exc}")

    _MC_V2_DEFAULT_PROMPT = (
        "Our Managed Services team follows the Cloudelligent A2Z structured methodology "
        "with four phases: Comprehensive Onboarding, Stabilize & Secure, Operate & Optimize, "
        "and Continuous Optimization & Modernization. "
        "Based on the Jira issues below, write a professional status report covering what has "
        "been completed and what remains in each phase. "
        "Be specific — reference actual task names and outcomes. "
        "The executive summary should be 3-5 sentences suitable for a customer stakeholder."
    )

    def _render_v2_prompt_editor():
        """Single-prompt editor for the MC V2 Audit (category='MC_V2')."""
        row = (
            db.query(AIAnalysisPrompt)
            .filter(AIAnalysisPrompt.category == 'MC_V2', AIAnalysisPrompt.is_active == True)
            .order_by(AIAnalysisPrompt.sequence_order)
            .first()
        )
        st.caption(
            "This prompt provides context and instructions to the AI when generating "
            "the Managed Services V2 Audit report for each customer. "
            "Edit and save to customise; the default reflects the A2Z methodology framework."
        )
        edited_text = st.text_area(
            "Audit Instructions",
            value=row.prompt_text if row else _MC_V2_DEFAULT_PROMPT,
            height=200,
            key="ai_prompt_MC_V2",
        )
        if st.button("Save MC V2 Audit Prompt", key="save_ai_MC_V2", type="primary"):
            try:
                db.query(AIAnalysisPrompt).filter(
                    AIAnalysisPrompt.category == 'MC_V2'
                ).delete()
                if edited_text.strip():
                    db.add(AIAnalysisPrompt(
                        category='MC_V2',
                        sequence_order=1,
                        prompt_text=edited_text.strip(),
                        is_active=True,
                    ))
                db.commit()
                st.success("MC V2 Audit prompt saved.")
                st.rerun()
            except Exception as exc:
                db.rollback()
                st.error(f"Failed to save prompt: {exc}")

    _FORECAST_DEFAULT_PROMPT = (
        "You are a resource planning analyst for a professional services firm.\n"
        "Review the forecast vs actual hours data below for the specified period.\n\n"
        "For each resource: classify their utilization status and provide a one-sentence observation.\n\n"
        "Status classifications:\n"
        "- On Track: 80-120% of forecast achieved\n"
        "- Over: >120% of forecast\n"
        "- Under: 50-80% of forecast\n"
        "- Critical Under: <50% of forecast (only when total_forecasted_hours > 10)\n"
        "- No Actuals: forecasted hours exist but zero logged\n"
        "- Unforecasted: hours logged with no forecast at all\n\n"
        "Pay particular attention to:\n"
        "- Resources with zero actuals despite significant forecasts (time submission or engagement issues)\n"
        "- Completely unforecasted resources logging full weeks (planning gaps)\n"
        "- Resources consistently below 50% of forecast (capacity or project issues)\n\n"
        "Provide 3-5 key observations and 2-3 actionable recommendations for the delivery management team.\n"
        "Return ONLY valid JSON matching the schema provided — no prose, no markdown."
    )

    def _render_forecast_prompt_editor():
        """Single-prompt editor for the Forecast Analysis (category='FORECAST')."""
        row = (
            db.query(AIAnalysisPrompt)
            .filter(AIAnalysisPrompt.category == 'FORECAST', AIAnalysisPrompt.is_active == True)
            .order_by(AIAnalysisPrompt.sequence_order)
            .first()
        )
        st.caption(
            "This prompt instructs the AI when analysing forecast vs actual hours across all resources. "
            "Edit and save to customise the analysis focus and output."
        )
        edited_text = st.text_area(
            "Analysis Instructions",
            value=row.prompt_text if row else _FORECAST_DEFAULT_PROMPT,
            height=280,
            key="ai_prompt_FORECAST",
        )
        if st.button("Save Forecast Analysis Prompt", key="save_ai_FORECAST", type="primary"):
            try:
                db.query(AIAnalysisPrompt).filter(
                    AIAnalysisPrompt.category == 'FORECAST'
                ).delete()
                if edited_text.strip():
                    db.add(AIAnalysisPrompt(
                        category='FORECAST',
                        sequence_order=1,
                        prompt_text=edited_text.strip(),
                        is_active=True,
                    ))
                db.commit()
                st.success("Forecast Analysis prompt saved.")
                st.rerun()
            except Exception as exc:
                db.rollback()
                st.error(f"Failed to save prompt: {exc}")

    with ai_ps_tab:
        _render_prompt_editor("PS")

    with ai_mc_tab:
        _render_prompt_editor("MC")

    with ai_v2_tab:
        _render_v2_prompt_editor()

    with ai_forecast_tab:
        _render_forecast_prompt_editor()

    for _cat, _tab in _extra_tab_objs.items():
        with _tab:
            _render_prompt_editor(_cat)

    # ── Week selector (shared by both run buttons) ─────────────────
    st.markdown("#### Run Analysis")
    _ai_current_monday = get_monday_of_week()
    _ai_week_options = {
        f"Last week ({(_ai_current_monday - timedelta(weeks=1)).strftime('%b %d')})": (_ai_current_monday - timedelta(weeks=1)).date(),
        f"2 weeks ago ({(_ai_current_monday - timedelta(weeks=2)).strftime('%b %d')})": (_ai_current_monday - timedelta(weeks=2)).date(),
        f"3 weeks ago ({(_ai_current_monday - timedelta(weeks=3)).strftime('%b %d')})": (_ai_current_monday - timedelta(weeks=3)).date(),
    }
    col_run1, col_run2 = st.columns([1, 3])
    with col_run1:
        run_week_label = st.selectbox(
            "Week to analyse",
            list(_ai_week_options.keys()),
            key="ai_run_weeks",
        )
    selected_week_start = _ai_week_options[run_week_label]

    col_btn1, col_btn2, col_btn3 = st.columns(3)

    with col_btn1:
        if st.button("Run AI Analysis Now", key="run_ai_analysis", type="primary"):
            if not aws_available:
                st.error("AWS credentials not available — cannot invoke Lambda.")
            else:
                with st.spinner("Invoking Lambda for AI analysis (this may take 1–2 minutes)..."):
                    try:
                        import boto3 as _b3, json as _json
                        _lc = _b3.client('lambda', region_name='us-east-1')
                        _resp = _lc.invoke(
                            FunctionName='production-clockify-import',
                            InvocationType='RequestResponse',
                            Payload=_json.dumps({
                                "mode": "analyze_project_health",
                                "week_start": str(selected_week_start),
                            })
                        )
                        _payload = _json.loads(_resp['Payload'].read())
                        if _resp.get('StatusCode') == 200 and 'FunctionError' not in _resp:
                            _body = _json.loads(_payload.get('body', '{}'))
                            _s = _body.get('analysis_summary', {})
                            ps_rows = _s.get('ps', {}).get('by_user_rows', '?')
                            mc_rows = _s.get('mc', {}).get('by_user_rows', '?')
                            st.success(
                                f"Analysis complete — PS: {ps_rows} user rows, MC: {mc_rows} user rows. "
                                "QuickSight SPICE refresh triggered."
                            )
                        else:
                            err = _payload.get('errorMessage') or _payload.get('body', 'Unknown error')
                            st.error(f"Lambda returned an error: {err}")
                    except Exception as exc:
                        st.error(f"Failed to invoke Lambda: {exc}")

    with col_btn2:
        if st.button("Run MC V2 Audit", key="run_mc_v2_audit", type="secondary"):
            if not aws_available:
                st.error("AWS credentials not available — cannot invoke Lambda.")
            else:
                with st.spinner("Invoking Lambda for MC V2 Audit (may take 2–4 minutes)..."):
                    try:
                        import boto3 as _b3, json as _json
                        _lc = _b3.client('lambda', region_name='us-east-1')
                        _resp = _lc.invoke(
                            FunctionName='production-clockify-import',
                            InvocationType='RequestResponse',
                            Payload=_json.dumps({
                                "mode": "mc_v2_audit",
                                "week_start": str(selected_week_start),
                            })
                        )
                        _payload = _json.loads(_resp['Payload'].read())
                        if _resp.get('StatusCode') == 200 and 'FunctionError' not in _resp:
                            _body = _json.loads(_payload.get('body', '{}'))
                            _audit = _body.get('audit_summary', {})
                            _processed = len([v for v in _audit.values() if 'error' not in v and 'skipped' not in v])
                            _errors = len([v for v in _audit.values() if 'error' in v])
                            st.success(
                                f"MC V2 Audit complete — {_processed} customers processed"
                                + (f", {_errors} errors" if _errors else "")
                                + ". QuickSight SPICE refresh triggered."
                            )
                            if _errors:
                                for cname, cval in _audit.items():
                                    if 'error' in cval:
                                        st.warning(f"{cname}: {cval['error']}")
                        else:
                            err = _payload.get('errorMessage') or _payload.get('body', 'Unknown error')
                            st.error(f"Lambda returned an error: {err}")
                    except Exception as exc:
                        st.error(f"Failed to invoke Lambda: {exc}")

    # ── Forecast Analysis run button ────────────────────────────────
    st.markdown("#### Run Forecast Analysis")
    _fcst_period_options = {
        "Last 2 weeks": 2,
        "Last 4 weeks (1 month)": 4,
        "Last 8 weeks (2 months)": 8,
    }
    col_fcst1, col_fcst2 = st.columns([1, 3])
    with col_fcst1:
        fcst_period_label = st.radio(
            "Analysis period",
            list(_fcst_period_options.keys()),
            index=1,
            key="fcst_period",
        )
    fcst_weeks_back = _fcst_period_options[fcst_period_label]

    if st.button("Run Forecast Analysis", key="run_forecast_analysis", type="primary"):
        if not aws_available:
            st.error("AWS credentials not available — cannot invoke Lambda.")
        else:
            with st.spinner(f"Running forecast analysis for {fcst_period_label.lower()} (may take 1–2 minutes)..."):
                try:
                    import boto3 as _b3, json as _json
                    _lc = _b3.client('lambda', region_name='us-east-1')
                    _ai_current_monday2 = get_monday_of_week()
                    _resp = _lc.invoke(
                        FunctionName='production-clockify-import',
                        InvocationType='RequestResponse',
                        Payload=_json.dumps({
                            "mode": "analyze_forecast",
                            "week_start": str(_ai_current_monday2.date()),
                            "weeks_back": fcst_weeks_back,
                        })
                    )
                    _payload = _json.loads(_resp['Payload'].read())
                    if _resp.get('StatusCode') == 200 and 'FunctionError' not in _resp:
                        _body = _json.loads(_payload.get('body', '{}'))
                        _s = _body.get('summary', {})
                        _users = _s.get('user_rows', '?')
                        _obs = _s.get('observations', '?')
                        _recs = _s.get('recommendations', '?')
                        st.success(
                            f"Forecast analysis complete — {_users} users analysed, "
                            f"{_obs} observations, {_recs} recommendations. "
                            "QuickSight SPICE refresh triggered."
                        )
                    else:
                        err = _payload.get('errorMessage') or _payload.get('body', 'Unknown error')
                        st.error(f"Lambda returned an error: {err}")
                except Exception as exc:
                    st.error(f"Failed to invoke Lambda: {exc}")

elif page == "Settings":
    st.header("⚙️ Settings")

    # Database Statistics
    st.subheader("📊 Database Statistics")

    user_count = db.query(func.count(ClockifyUser.user_id)).scalar()
    active_user_count = db.query(func.count(ClockifyUser.user_id)).filter(
        ClockifyUser.status == 'active'
    ).scalar()
    entry_count = db.query(func.count(ClockifyTimeEntry.entry_id)).scalar()
    project_count = db.query(func.count(func.distinct(ClockifyTimeEntry.project_name))).scalar()

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Users", user_count)
    with col2:
        st.metric("Active Users", active_user_count)
    with col3:
        st.metric("Total Projects", project_count)
    with col4:
        st.metric("Time Entries", entry_count)

    st.divider()

    # User Management Section — manages who can log in to this app
    st.subheader("👤 User Management")
    st.caption("Manage login accounts for this application.")

    import bcrypt as _bcrypt

    # ── Current users table ───────────────────────────────────────
    app_users = db.query(AppUser).order_by(AppUser.username).all()
    if app_users:
        user_rows = [
            {
                "Username": u.username,
                "Display Name": u.display_name,
                "Status": "Active" if u.is_active else "Inactive",
                "Created": u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
            }
            for u in app_users
        ]
        st.dataframe(pd.DataFrame(user_rows), use_container_width=True, hide_index=True)
    else:
        st.info("No app users found.")

    with st.expander("Add New User", expanded=False):
        st.markdown("Create a new login account for this application.")

        with st.form("add_user_form"):
            new_username = st.text_input("Username *", placeholder="jdoe")
            new_display_name = st.text_input("Display Name *", placeholder="Jane Doe")
            new_password = st.text_input("Password *", type="password")
            new_password_confirm = st.text_input("Confirm Password *", type="password")

            submitted = st.form_submit_button("Add User", type="primary")

            if submitted:
                if not new_username or not new_display_name or not new_password:
                    st.error("Username, display name, and password are all required.")
                elif new_password != new_password_confirm:
                    st.error("Passwords do not match.")
                else:
                    existing = db.query(AppUser).filter(
                        func.lower(AppUser.username) == func.lower(new_username)
                    ).first()
                    if existing:
                        st.error(f"Username '{new_username}' is already taken.")
                    else:
                        try:
                            pw_hash = _bcrypt.hashpw(
                                new_password.encode(), _bcrypt.gensalt(12)
                            ).decode()
                            db.add(AppUser(
                                username=new_username.strip(),
                                display_name=new_display_name.strip(),
                                password_hash=pw_hash,
                                is_active=True,
                            ))
                            db.commit()
                            st.success(f"User '{new_username}' created successfully.")
                            st.rerun()
                        except Exception as e:
                            db.rollback()
                            st.error(f"Failed to create user: {e}")

    # Edit existing app users
    with st.expander("Edit User", expanded=False):
        edit_user_map = {u.username: u for u in app_users}
        selected_username = st.selectbox(
            "Select user to edit",
            options=[""] + list(edit_user_map.keys()),
            key="edit_app_user_select",
        )

        if selected_username:
            eu = edit_user_map[selected_username]

            with st.form("edit_app_user_form"):
                eu_display = st.text_input("Display Name", value=eu.display_name)
                eu_status = st.selectbox(
                    "Status",
                    options=["Active", "Inactive"],
                    index=0 if eu.is_active else 1,
                )
                st.markdown("**Reset Password** — leave blank to keep current password")
                eu_new_pw = st.text_input("New Password", type="password", key="eu_pw")
                eu_new_pw2 = st.text_input("Confirm New Password", type="password", key="eu_pw2")

                update_submitted = st.form_submit_button("Save Changes", type="primary")

                if update_submitted:
                    if not eu_display.strip():
                        st.error("Display name cannot be blank.")
                    elif eu_new_pw and eu_new_pw != eu_new_pw2:
                        st.error("Passwords do not match.")
                    else:
                        active_count = db.query(AppUser).filter(AppUser.is_active == True).count()
                        going_inactive = (eu_status == "Inactive" and eu.is_active)
                        if going_inactive and active_count <= 1:
                            st.error("Cannot deactivate the last active user.")
                        else:
                            try:
                                eu.display_name = eu_display.strip()
                                eu.is_active = (eu_status == "Active")
                                if eu_new_pw:
                                    eu.password_hash = _bcrypt.hashpw(
                                        eu_new_pw.encode(), _bcrypt.gensalt(12)
                                    ).decode()
                                eu.updated_at = datetime.now()
                                db.commit()
                                st.success(f"User '{selected_username}' updated.")
                                st.rerun()
                            except Exception as e:
                                db.rollback()
                                st.error(f"Failed to update user: {e}")

    # Delete app users
    with st.expander("Remove User", expanded=False):
        delete_user_map = {u.username: u for u in app_users}
        del_username = st.selectbox(
            "Select user to remove",
            options=[""] + list(delete_user_map.keys()),
            key="del_app_user_select",
        )
        if del_username:
            du = delete_user_map[del_username]
            st.warning(f"This will permanently delete the account for **{du.display_name}** ({du.username}).")
            if st.button("Delete User", type="primary", key="confirm_del_user"):
                active_count = db.query(AppUser).filter(AppUser.is_active == True).count()
                if du.is_active and active_count <= 1:
                    st.error("Cannot delete the last active user.")
                else:
                    try:
                        db.delete(du)
                        db.commit()
                        st.success(f"User '{del_username}' removed.")
                        st.rerun()
                    except Exception as e:
                        db.rollback()
                        st.error(f"Failed to delete user: {e}")

    st.divider()

    # ── Notification Recipients ───────────────────────────────────────────
    st.subheader("📧 Notification Recipients")
    st.caption("Manage who receives the Monday import status email.")

    REPORT_RUN_LABELS = {
        'morning':   'Morning import only (9am)',
        'noon':      'Noon import only (12pm)',
        'afternoon': 'Afternoon only',
        'both':      'Both (9am + noon)',
        'all':       'All reports',
    }

    try:
        with engine.connect() as _nr_conn:
            _nr_rows = _nr_conn.execute(text(
                "SELECT id, email, display_name, report_run, is_active "
                "FROM compliance_report_recipients ORDER BY email"
            )).fetchall()
        _nr_load_err = None
    except Exception as _nr_e:
        st.error(f"Could not load recipients: {_nr_e}")
        _nr_rows = []
        _nr_load_err = _nr_e

    _nr_active_count = sum(1 for r in _nr_rows if r[4])
    st.caption(f"{_nr_active_count} active recipient{'s' if _nr_active_count != 1 else ''}")

    if _nr_rows:
        _hc = st.columns([2, 2.5, 2, 0.8, 0.7])
        _hc[0].markdown("**Name**")
        _hc[1].markdown("**Email**")
        _hc[2].markdown("**Receives**")
        _hc[3].markdown("**Active**")
        _hc[4].markdown("**Remove**")
        st.divider()

        for _nr_id, _nr_email, _nr_name, _nr_run, _nr_is_active in _nr_rows:
            _rc = st.columns([2, 2.5, 2, 0.8, 0.7])
            _rc[0].text(_nr_name or "—")
            _rc[1].text(_nr_email)
            _rc[2].text(REPORT_RUN_LABELS.get(_nr_run or '', _nr_run or "—"))
            with _rc[3]:
                _new_active = st.toggle(
                    "",
                    value=bool(_nr_is_active),
                    key=f"nr_toggle_{_nr_id}",
                    label_visibility="collapsed"
                )
                if _new_active != bool(_nr_is_active):
                    try:
                        with engine.begin() as _tc:
                            _tc.execute(
                                text("UPDATE compliance_report_recipients SET is_active=:v WHERE id=:id"),
                                {"v": _new_active, "id": _nr_id}
                            )
                        st.rerun()
                    except Exception as _te:
                        st.error(f"Update failed: {_te}")
            with _rc[4]:
                if st.button("✕", key=f"nr_del_{_nr_id}", help="Remove recipient"):
                    try:
                        with engine.begin() as _dc:
                            _dc.execute(
                                text("DELETE FROM compliance_report_recipients WHERE id=:id"),
                                {"id": _nr_id}
                            )
                        st.rerun()
                    except Exception as _de:
                        st.error(f"Remove failed: {_de}")
    else:
        if _nr_load_err is None:
            st.info("No recipients configured yet.")

    with st.expander("➕ Add Recipient", expanded=False):
        with st.form("settings_add_recipient_form"):
            _add_name = st.text_input("Display Name (optional)", placeholder="Chris Xenos")
            _add_email = st.text_input("Email *", placeholder="user@cloudelligent.com")
            _add_run = st.selectbox(
                "Receives",
                options=list(REPORT_RUN_LABELS.keys()),
                format_func=lambda k: REPORT_RUN_LABELS[k],
            )
            _add_submitted = st.form_submit_button("Add Recipient", type="primary")
            if _add_submitted:
                _email_val = (_add_email or "").strip()
                if not _email_val or "@" not in _email_val or "." not in _email_val.split("@")[-1]:
                    st.error("Please enter a valid email address.")
                else:
                    try:
                        with engine.begin() as _ac:
                            _ac.execute(text(
                                "INSERT INTO compliance_report_recipients "
                                "(email, display_name, report_run) VALUES (:e, :n, :r) "
                                "ON CONFLICT (email) DO UPDATE SET "
                                "display_name=EXCLUDED.display_name, "
                                "report_run=EXCLUDED.report_run, is_active=TRUE"
                            ), {"e": _email_val, "n": (_add_name or "").strip() or None, "r": _add_run})
                        st.toast("Recipient added", icon="✅")
                        st.rerun()
                    except Exception as _ae:
                        st.error(f"Failed to add recipient: {_ae}")

    st.divider()

    # Date range settings
    with st.expander("📅 Default Date Range Settings", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            weeks_back = st.number_input(
                "Default Weeks Back",
                min_value=0,
                max_value=52,
                value=st.session_state.weeks_back,
                help="Default number of weeks to look back in reports"
            )
            st.session_state.weeks_back = weeks_back
        
        with col2:
            weeks_forward = st.number_input(
                "Default Weeks Forward",
                min_value=0,
                max_value=52,
                value=st.session_state.weeks_forward,
                help="Default number of weeks to look forward for forecasts"
            )
            st.session_state.weeks_forward = weeks_forward
        
        st.caption(f"💡 Default date range: {weeks_back} weeks back to {weeks_forward} weeks forward")
    
    st.divider()
    
    # Custom Fields Summary
    with st.expander("📋 Custom Fields Configuration", expanded=False):
        st.info("The following custom fields are configured in Clockify:")
    
        custom_fields_df = pd.DataFrame([
            {"Field Name": "Practice Alignment", "Type": "Text", "Purpose": "Service Line classification"},
            {"Field Name": "Skill Area", "Type": "Text", "Purpose": "Technical expertise area"},
            {"Field Name": "POD Assignment", "Type": "Text", "Purpose": "Team/pod membership"},
            {"Field Name": "Cloudelligent Title", "Type": "Text", "Purpose": "Job title"},
            {"Field Name": "Location", "Type": "Text", "Purpose": "Onshore/Offshore"},
            {"Field Name": "Employment Designation", "Type": "Text", "Purpose": "FTE/Contractor status"}
        ])
    
        st.dataframe(custom_fields_df, use_container_width=True)
    
    st.divider()
    
    # Practice Alignment Distribution
    with st.expander("👥 Practice Alignment Distribution", expanded=False):
        practice_counts = db.query(
            ClockifyUser.practice_alignment,
            func.count(ClockifyUser.user_id).label("count")
        ).filter(
            ClockifyUser.status == 'active'
        ).group_by(
            ClockifyUser.practice_alignment
        ).all()

        if practice_counts:
            df = pd.DataFrame([
                {"Practice Alignment": pa or "Unassigned", "Active Users": count}
                for pa, count in practice_counts
            ])
            st.dataframe(df, use_container_width=True)
    
    st.divider()
    st.subheader('🏢 Line of Business Mapping')
    st.caption('Maps Practice Alignment values to Lines of Business. Changes take effect on the next data refresh.')

    try:
        _lob_rows = db.execute(text(
            'SELECT id, practice_alignment, line_of_business FROM lob_practice_mapping ORDER BY line_of_business, practice_alignment'
        )).fetchall()
    except Exception as _lob_load_err:
        st.error(f'Could not load LoB mapping: {_lob_load_err}')
        _lob_rows = []

    if _lob_rows:
        _lob_df = pd.DataFrame(_lob_rows, columns=['id', 'Practice Alignment', 'Line of Business'])
        _lob_options = ['Professional Services', 'Managed Cloud', 'Managed IT', 'FINOPs', 'Product', 'Internal']
        _edited_lob = st.data_editor(
            _lob_df[['Practice Alignment', 'Line of Business']],
            use_container_width=True,
            hide_index=True,
            column_config={
                'Practice Alignment': st.column_config.TextColumn('Practice Alignment', disabled=True),
                'Line of Business': st.column_config.SelectboxColumn('Line of Business', options=_lob_options, required=True)
            },
            key='lob_mapping_editor'
        )
        if st.button('💾 Save LoB Mapping', type='primary', key='save_lob_mapping'):
            try:
                for i, row in _edited_lob.iterrows():
                    orig_pa = _lob_df.iloc[i]['Practice Alignment']
                    db.execute(text(
                        'UPDATE lob_practice_mapping SET line_of_business = :lob, updated_at = NOW() WHERE practice_alignment = :pa'
                    ), {'lob': row['Line of Business'], 'pa': orig_pa})
                db.commit()
                st.success('LoB mapping saved.')
                st.rerun()
            except Exception as e:
                db.rollback()
                st.error(f'Save failed: {e}')
    else:
        st.info('No LoB mappings configured. Use the form below to add entries.')

    with st.expander('Add New Practice Alignment', expanded=False):
        with st.form('add_lob_mapping_form'):
            _new_pa = st.text_input('Practice Alignment', placeholder='e.g. Cloud Foundations')
            _new_lob = st.selectbox('Line of Business', ['Professional Services', 'Managed Cloud', 'Managed IT', 'FINOPs', 'Product', 'Internal'])
            if st.form_submit_button('Add', type='primary'):
                try:
                    db.execute(text(
                        'INSERT INTO lob_practice_mapping (practice_alignment, line_of_business) VALUES (:pa, :lob) ON CONFLICT (practice_alignment) DO UPDATE SET line_of_business = :lob, updated_at = NOW()'
                    ), {'pa': _new_pa.strip(), 'lob': _new_lob})
                    db.commit()
                    st.success(f'Added: {_new_pa} → {_new_lob}')
                    st.rerun()
                except Exception as e:
                    db.rollback()
                    st.error(f'Failed: {e}')

    st.divider()

    st.caption("⚠️ SES domain verification required before emails will send. Check with: `aws sesv2 get-email-identity --email-identity cloudelligent.com`")
    
    st.divider()
    
    # Data freshness
    with st.expander("🔄 Data Freshness", expanded=False):
        latest_entry = db.query(func.max(ClockifyTimeEntry.synced_at)).scalar()
        
        if latest_entry:
            st.info(f"Last data sync: {latest_entry.strftime('%Y-%m-%d %H:%M:%S')}")
        else:
            st.warning("No data synced yet. Go to 'Data Import' to import data.")

elif page == "Resource Forecast":
    st.header("📈 Resource Forecast")

    tab_upload, tab_view, tab_history, tab_capacity = st.tabs([
        "📤 Upload", "📊 View", "📋 History", "🗂️ Capacity"
    ])

    with tab_upload:
        st.subheader("Upload Weekly Forecast Template")
        st.markdown("""
        **Supported Format:** Upload your weekly forecasting Excel file with the standard template format:
        - Row 1: Week date ranges (e.g., "16th-20th Dec")
        - Row 2: Week labels (Week1, Week2, etc.)
        - Row 4: Headers (Client, Comments, Type, PM, Stage, User, Plan/Actual columns)
        - Data rows: One row per user per project with Plan hours for each week
        """)

        st.markdown("#### Download Template")
        col_dl1, col_dl2 = st.columns([1, 3])
        with col_dl1:
            def generate_template_excel():
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = "Forecast"
                current_monday = get_monday_of_week()
                weeks = []
                for i in range(12):
                    week_start = current_monday + timedelta(weeks=i)
                    week_end = week_start + timedelta(days=4)
                    weeks.append({
                        'start': week_start, 'end': week_end,
                        'date_range': f"{week_start.strftime('%d')}-{week_end.strftime('%d %b')}",
                        'label': f"Week{i+1}"
                    })
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                for i, week in enumerate(weeks):
                    ws.cell(row=1, column=9 + i, value=week['date_range'])
                    ws.cell(row=2, column=9 + i, value=week['label'])
                headers = ['', 'Client', 'Project', 'Comments', 'Type', 'PM', 'Stage', 'User'] + ['Plan'] * len(weeks)
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                active_users = db.query(ClockifyUser).filter(ClockifyUser.status == 'active').order_by(ClockifyUser.name).all()
                row_num = 5
                for client in ['Sample Client 1', 'Sample Client 2']:
                    for user in active_users[:5]:
                        ws.cell(row=row_num, column=2, value=client)
                        ws.cell(row=row_num, column=3, value=f"{client} Project")
                        ws.cell(row=row_num, column=5, value="Migration")
                        ws.cell(row=row_num, column=7, value="Build and Implement")
                        ws.cell(row=row_num, column=8, value=user.name)
                        for i in range(len(weeks)):
                            ws.cell(row=row_num, column=9 + i, value=0)
                        row_num += 1
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 18
                ws.column_dimensions['H'].width = 20
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()

            template_data = generate_template_excel()
            st.download_button(
                label="📥 Download Template",
                data=template_data,
                file_name=f"forecast_template_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        with col_dl2:
            st.caption("Download a blank template pre-filled with active users to enter your forecasts.")

        st.divider()
        uploaded_file = st.file_uploader(
            "Upload weekly forecast Excel file",
            type=['xlsx', 'xls'],
            help="Upload your weekly forecasting template"
        )

        if uploaded_file is not None:
            try:
                from src.integrations.forecast_parser import parse_forecast_template
                with st.spinner("Parsing forecast template..."):
                    forecasts, template_week_dates = parse_forecast_template(uploaded_file)

                if not forecasts:
                    st.warning("No forecast data found in the file. Please check the format.")
                else:
                    st.subheader("Preview")
                    preview_df = pd.DataFrame(forecasts[:50])
                    if 'week_start_date' in preview_df.columns:
                        preview_df['week_start_date'] = pd.to_datetime(preview_df['week_start_date']).dt.strftime('%Y-%m-%d')
                    if 'week_end_date' in preview_df.columns:
                        preview_df['week_end_date'] = pd.to_datetime(preview_df['week_end_date']).dt.strftime('%Y-%m-%d')
                    st.dataframe(preview_df, use_container_width=True)
                    st.caption(f"Showing first 50 of {len(forecasts)} forecast entries")

                    st.subheader("Summary")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Entries", len(forecasts))
                    with col2:
                        st.metric("Unique Users", len(set(f['user_name'] for f in forecasts)))
                    with col3:
                        st.metric("Unique Clients", len(set(f['client_name'] for f in forecasts)))
                    with col4:
                        st.metric("Total Hours", f"{sum(f.get('forecasted_hours', 0) for f in forecasts):.0f}")

                    if st.button("📥 Import Forecasts", type="primary"):
                        with st.spinner("Importing forecasts..."):
                            imported, updated, skipped_users = import_forecasts_from_template(db, forecasts, template_week_dates)
                            st.success(f"Imported {imported} new forecasts, updated {updated} existing")
                            if skipped_users:
                                st.warning(f"⚠️ {len(skipped_users)} user(s) not found in Clockify — their rows were **rejected**: {', '.join(sorted(skipped_users))}")

                        forecast_dataset_ids = [
                            "8900f5dc-687e-4d5b-9f91-5efd0cd1daed",
                            "3bdc816d-4df6-4db7-b3e6-64e230f28f14",
                            "42098a5b-a94f-41d5-8300-396f1fec66bf",
                            "fc56c886-f0d2-4935-8b32-f0862325d3f0",
                            "ps-project-status-view",
                            "free-agent-availability",
                            "data-freshness",
                        ]
                        with st.spinner("Refreshing QuickSight datasets..."):
                            try:
                                import boto3
                                import json as _json
                                lambda_client = boto3.client('lambda', region_name='us-east-1')
                                response = lambda_client.invoke(
                                    FunctionName='production-clockify-import',
                                    InvocationType='RequestResponse',
                                    Payload=_json.dumps({
                                        "mode": "refresh_quicksight_only",
                                        "quicksight_dataset_ids": forecast_dataset_ids
                                    })
                                )
                                if response.get('StatusCode') == 200 and 'FunctionError' not in response:
                                    st.success("Triggered SPICE refresh for 7 QuickSight datasets")
                                else:
                                    st.warning("QuickSight refresh returned an error. Datasets may need manual refresh.")
                            except Exception as e:
                                st.warning(f"Could not auto-refresh QuickSight datasets: {e}")
                        st.rerun()

            except Exception as e:
                st.error(f"Failed to parse file: {str(e)}")
                import traceback
                with st.expander("View Error Details"):
                    st.code(traceback.format_exc())

    with tab_upload:
        st.subheader("✏️ Manual Entry")
        st.info("Select a client, project, and staff members to enter forecasted hours for the next 16 weeks.")

        users = db.query(ClockifyUser).filter(ClockifyUser.status == 'active').order_by(ClockifyUser.name).all()
        projects = db.query(ClockifyProject).filter(ClockifyProject.archived == False).order_by(ClockifyProject.client_name, ClockifyProject.name).all()
        user_options = {u.name: u for u in users}
        clients = sorted(set(p.client_name for p in projects if p.client_name))

        current_monday = get_monday_of_week()
        weeks = []
        for i in range(16):
            week_start = current_monday + timedelta(weeks=i)
            week_end = week_start + timedelta(days=6)
            weeks.append({
                'start': week_start.date(), 'end': week_end.date(),
                'label': week_start.strftime('%b %d'),
                'full_label': f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d')}"
            })

        selected_client = st.selectbox("Select Client", options=[""] + clients, key="manual_client_select")
        filtered_projects = [p for p in projects if p.client_name == selected_client] if selected_client else projects
        project_options = {p.name: p for p in filtered_projects}

        col1, col2 = st.columns(2)
        with col1:
            selected_project_name = st.selectbox("Select Project", options=[""] + list(project_options.keys()), key="manual_project_select", disabled=not selected_client)
        with col2:
            selected_staff = st.multiselect("Select Staff Members", options=list(user_options.keys()), key="manual_staff_select")

        if selected_project_name and selected_staff:
            project = project_options[selected_project_name]
            st.divider()
            st.markdown(f"**Project:** {project.name} ({project.client_name or 'No Client'})")
            week_cols = [w['label'] for w in weeks]

            existing_forecasts = db.query(PSResourceForecast).filter(
                PSResourceForecast.project_name == project.name,
                PSResourceForecast.user_name.in_(selected_staff),
                PSResourceForecast.week_start_date >= weeks[0]['start'],
                PSResourceForecast.week_start_date <= weeks[-1]['start']
            ).all()
            forecast_lookup = {(f.user_name, f.week_start_date): f.forecasted_hours for f in existing_forecasts}

            grid_data = []
            for staff_name in selected_staff:
                row = {'Staff': staff_name}
                for week in weeks:
                    row[week['label']] = forecast_lookup.get((staff_name, week['start']), 0.0)
                grid_data.append(row)

            df = pd.DataFrame(grid_data)
            st.caption("Week columns: " + " | ".join([f"{w['label']}: {w['full_label']}" for w in weeks[:4]]) + " ...")
            edited_df = st.data_editor(
                df, use_container_width=True, num_rows="fixed",
                column_config={
                    "Staff": st.column_config.TextColumn("Staff", disabled=True, width="medium"),
                    **{w['label']: st.column_config.NumberColumn(w['label'], min_value=0.0, max_value=80.0, step=0.5, format="%.1f", width="small") for w in weeks}
                },
                key="forecast_grid"
            )
            st.divider()
            total_all = edited_df[week_cols].sum().sum()
            st.markdown(f"**Total Hours:** {total_all:.1f} hours across all staff and weeks")

            col_save, _ = st.columns([1, 3])
            with col_save:
                if st.button("💾 Save All Forecasts", type="primary", key="save_forecasts"):
                    try:
                        saved_count = 0
                        updated_count = 0
                        for _, row in edited_df.iterrows():
                            staff_name = row['Staff']
                            user = user_options.get(staff_name)
                            if not user:
                                continue
                            for week in weeks:
                                hours = row[week['label']]
                                if pd.isna(hours):
                                    hours = 0.0
                                existing = db.query(PSResourceForecast).filter(
                                    PSResourceForecast.week_start_date == week['start'],
                                    PSResourceForecast.user_name == staff_name,
                                    PSResourceForecast.project_name == project.name
                                ).first()
                                if existing:
                                    if existing.forecasted_hours != hours:
                                        existing.forecasted_hours = hours
                                        existing.updated_at = datetime.now()
                                        updated_count += 1
                                elif hours > 0:
                                    db.add(PSResourceForecast(
                                        week_start_date=week['start'], week_end_date=week['end'],
                                        clockify_user_id=user.clockify_user_id, user_name=user.name,
                                        location=user.location, employment_designation=user.employment_designation,
                                        project_name=project.name, clockify_project_id=project.clockify_project_id,
                                        client_name=project.client_name or project.name,
                                        practice_area=user.practice_alignment, forecasted_hours=hours
                                    ))
                                    saved_count += 1
                        db.commit()
                        st.success(f"✅ Saved {saved_count} new forecasts, updated {updated_count} existing")
                        st.rerun()
                    except Exception as e:
                        db.rollback()
                        st.error(f"Failed to save: {str(e)}")
        elif selected_project_name and not selected_staff:
            st.info("👆 Select one or more staff members to enter forecasts.")
        else:
            st.info("👆 Select a project to begin entering forecasts.")

    with tab_view:
        st.subheader("View Existing Forecasts")
        col1, col2, col3 = st.columns(3)
        with col1:
            view_weeks = st.selectbox("Time Range", ["All Weeks", "Next 4 Weeks", "Next 8 Weeks", "Next 12 Weeks", "All Future"])
        with col2:
            filter_user = st.multiselect("Filter by Resource", options=[u.name for u in db.query(ClockifyUser).filter(ClockifyUser.status == 'active').all()])
        with col3:
            filter_client = st.multiselect("Filter by Client", options=[c[0] for c in db.query(PSResourceForecast.client_name).distinct().all() if c[0]])

        col4, col5, col6 = st.columns(3)
        with col4:
            filter_pm = st.multiselect("Filter by PM", options=[p[0] for p in db.query(PSResourceForecast.pm_name).distinct().all() if p[0]])
        with col5:
            filter_type = st.multiselect("Filter by Type", options=[t[0] for t in db.query(PSResourceForecast.project_type).distinct().all() if t[0]])
        with col6:
            filter_stage = st.multiselect("Filter by Stage", options=[s[0] for s in db.query(PSResourceForecast.stage).distinct().all() if s[0]])

        current_monday = get_monday_of_week()
        query = db.query(PSResourceForecast)
        if view_weeks == "Next 4 Weeks":
            query = query.filter(PSResourceForecast.week_start_date >= current_monday.date(), PSResourceForecast.week_start_date < (current_monday + timedelta(weeks=4)).date())
        elif view_weeks == "Next 8 Weeks":
            query = query.filter(PSResourceForecast.week_start_date >= current_monday.date(), PSResourceForecast.week_start_date < (current_monday + timedelta(weeks=8)).date())
        elif view_weeks == "Next 12 Weeks":
            query = query.filter(PSResourceForecast.week_start_date >= current_monday.date(), PSResourceForecast.week_start_date < (current_monday + timedelta(weeks=12)).date())
        elif view_weeks == "All Future":
            query = query.filter(PSResourceForecast.week_start_date >= current_monday.date())
        if filter_user:
            query = query.filter(PSResourceForecast.user_name.in_(filter_user))
        if filter_client:
            query = query.filter(PSResourceForecast.client_name.in_(filter_client))
        if filter_pm:
            query = query.filter(PSResourceForecast.pm_name.in_(filter_pm))
        if filter_type:
            query = query.filter(PSResourceForecast.project_type.in_(filter_type))
        if filter_stage:
            query = query.filter(PSResourceForecast.stage.in_(filter_stage))

        forecasts_view = query.order_by(PSResourceForecast.week_start_date, PSResourceForecast.user_name).all()

        if forecasts_view:
            df = pd.DataFrame([{
                "Week Start": f.week_start_date, "Week End": f.week_end_date,
                "Resource": f.user_name, "Client": f.client_name or "",
                "Project": f.project_name or f.client_name or "",
                "Type": f.project_type or "", "PM": f.pm_name or "",
                "Stage": f.stage or "", "Hours": f.forecasted_hours,
                "Comments": f.comments or ""
            } for f in forecasts_view])

            col_toggle, col_download = st.columns([3, 1])
            with col_toggle:
                view_type = st.radio("View Type", ["Pivot Table", "List View"], horizontal=True)

            def _generate_forecast_download(df):
                weeks_df = df[['Week Start', 'Week End']].drop_duplicates().sort_values('Week Start')
                week_cols = [(w['Week Start'], f"{w['Week Start'].strftime('%d')}-{w['Week End'].strftime('%d %b')} Plan") for _, w in weeks_df.iterrows()]
                pivot = df.pivot_table(index=['Client', 'Project', 'Type', 'PM', 'Stage', 'Resource', 'Comments'], columns='Week Start', values='Hours', aggfunc='sum', fill_value=0).reset_index()
                output_rows = []
                for _, row in pivot.iterrows():
                    out = {'Client': row['Client'], 'Comments': row['Comments'], 'Type': row['Type'], 'PM': row['PM'], 'Stage': row['Stage'], 'User': row['Resource'], 'Total Plan': 0, 'Total Actual': 0}
                    total = 0
                    for ws, col_name in week_cols:
                        h = row.get(ws, 0) if ws in row.index else 0
                        out[col_name] = h
                        out[col_name.replace(' Plan', ' Actual')] = 0
                        total += h
                    out['Total Plan'] = total
                    output_rows.append(out)
                ordered = ['Client', 'Comments', 'Type', 'PM', 'Stage', 'User', 'Total Plan', 'Total Actual']
                for ws, col_name in week_cols:
                    ordered += [col_name, col_name.replace(' Plan', ' Actual')]
                out_df = pd.DataFrame(output_rows)
                out_df = out_df[[c for c in ordered if c in out_df.columns]]
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    out_df.to_excel(writer, sheet_name='Forecast', index=False)
                buf.seek(0)
                return buf

            with col_download:
                st.download_button(
                    label="📥 Download Forecast",
                    data=_generate_forecast_download(df),
                    file_name=f"forecast_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            if view_type == "Pivot Table":
                weeks_df = df[['Week Start', 'Week End']].drop_duplicates().sort_values('Week Start')
                week_cols_sorted = [(row['Week Start'], f"{row['Week Start'].strftime('%d')}-{row['Week End'].strftime('%d %b')}") for _, row in weeks_df.iterrows()]
                df['Date Range'] = df.apply(lambda r: f"{r['Week Start'].strftime('%d')}-{r['Week End'].strftime('%d %b')}", axis=1)
                pivot_df = df.pivot_table(index=['Client', 'Project', 'PM', 'Resource'], columns='Date Range', values='Hours', aggfunc='sum', fill_value=0)
                week_order = [c[1] for c in week_cols_sorted]
                pivot_df = pivot_df.reindex(columns=[c for c in week_order if c in pivot_df.columns])
                pivot_df['Total'] = pivot_df.sum(axis=1)
                st.dataframe(pivot_df.reset_index(), use_container_width=False, height=500)
                st.divider()
                st.subheader("Weekly Totals")
                weekly_totals = df.groupby('Week Start')['Hours'].sum().reset_index().sort_values('Week Start')
                cols = st.columns(min(len(weekly_totals), 12))
                for i, (_, row) in enumerate(weekly_totals.iterrows()):
                    if i < 12:
                        with cols[i]:
                            st.metric(row['Week Start'].strftime('%b %d'), f"{row['Hours']:.0f}h")
            else:
                df_display = df.copy()
                df_display['Week'] = df_display['Week Start'].apply(lambda x: x.strftime('%b %d'))
                st.dataframe(df_display[['Week', 'Resource', 'Client', 'Project', 'Type', 'PM', 'Stage', 'Hours', 'Comments']], use_container_width=True, height=400)

            st.divider()
            col1, col2, col3 = st.columns(3)
            with col1: st.metric("Total Forecasted Hours", f"{df['Hours'].sum():.1f}")
            with col2: st.metric("Resources", df['Resource'].nunique())
            with col3: st.metric("Clients", df['Client'].nunique())
        else:
            st.info("No forecasts found. Use the 'Upload Excel' or 'Manual Entry' tabs to add forecasts.")

    with tab_history:
        st.subheader("Forecast History")
        st.info("View how forecasts have changed over time. Each upload creates a snapshot of the previous forecast before replacing it.")
        from sqlalchemy import text as sa_text
        try:
            snapshots = db.execute(sa_text("""
                SELECT snapshot_id, archived_at,
                       COUNT(*) AS record_count,
                       COUNT(DISTINCT week_start_date) AS weeks,
                       COUNT(DISTINCT user_name) AS users,
                       SUM(forecasted_hours) AS total_hours
                FROM ps_resource_forecast_history
                GROUP BY snapshot_id, archived_at
                ORDER BY archived_at DESC
                LIMIT 20
            """)).fetchall()
        except Exception:
            snapshots = []

        if not snapshots:
            st.info("No forecast history yet. History is created automatically when you upload a new forecast that replaces existing data.")
        else:
            snapshot_options = {
                f"{s[1].strftime('%Y-%m-%d %H:%M')} ({s[2]} records, {s[4]} users, {s[5]:.0f} hrs)": s[0]
                for s in snapshots
            }
            hist_tab1, hist_tab2 = st.tabs(["Current vs Previous", "Browse Snapshot"])
            with hist_tab1:
                st.markdown("Shows what changed between the current forecast and the most recent previous version.")
                try:
                    comparison = db.execute(sa_text("SELECT * FROM vw_forecast_version_comparison ORDER BY week_start_date, user_name, client_name")).fetchall()
                except Exception:
                    comparison = []
                if comparison:
                    comp_df = pd.DataFrame(comparison)
                    comp_df.columns = ["Week Start", "Resource", "Client", "Project", "PM", "Type", "Stage", "Current Hours", "Previous Hours", "Change", "Change Type", "Previous Snapshot Date", "Previous Snapshot ID"]
                    h_col1, h_col2 = st.columns(2)
                    with h_col1:
                        h_filter_change = st.multiselect("Change Type", options=["New", "Removed", "Increased", "Decreased", "Unchanged"], default=["New", "Removed", "Increased", "Decreased"], key="hist_change_filter")
                    with h_col2:
                        h_filter_client = st.multiselect("Client", options=sorted(comp_df["Client"].dropna().unique()), key="hist_client_filter")
                    filtered = comp_df.copy()
                    if h_filter_change:
                        filtered = filtered[filtered["Change Type"].isin(h_filter_change)]
                    if h_filter_client:
                        filtered = filtered[filtered["Client"].isin(h_filter_client)]
                    c1, c2, c3, c4 = st.columns(4)
                    with c1: st.metric("New", len(filtered[filtered["Change Type"] == "New"]))
                    with c2: st.metric("Removed", len(filtered[filtered["Change Type"] == "Removed"]))
                    with c3: st.metric("Increased", len(filtered[filtered["Change Type"] == "Increased"]))
                    with c4: st.metric("Decreased", len(filtered[filtered["Change Type"] == "Decreased"]))
                    st.dataframe(filtered[["Week Start", "Resource", "Client", "Project", "Current Hours", "Previous Hours", "Change", "Change Type"]], use_container_width=True, height=500)
                else:
                    st.info("No comparison data available. Upload a forecast twice to see changes.")
            with hist_tab2:
                selected_label = st.selectbox("Select Snapshot", options=list(snapshot_options.keys()))
                selected_id = snapshot_options[selected_label]
                snapshot_data = db.execute(sa_text("""
                    SELECT week_start_date, user_name, client_name, project_name,
                           pm_name, project_type, stage, forecasted_hours
                    FROM ps_resource_forecast_history
                    WHERE snapshot_id = :sid
                    ORDER BY week_start_date, user_name, client_name
                """), {"sid": selected_id}).fetchall()
                if snapshot_data:
                    snap_df = pd.DataFrame(snapshot_data, columns=["Week Start", "Resource", "Client", "Project", "PM", "Type", "Stage", "Hours"])
                    st.dataframe(snap_df, use_container_width=True, height=500)
                    st.caption(f"Snapshot {selected_id}: {len(snapshot_data)} records, {snap_df['Hours'].sum():.0f} total hours")

    with tab_capacity:
        st.subheader("Resource Capacity Plan")
        _cap_monday = get_monday_of_week()

        # ── Filter controls — PS-only scope (practice_area IN ('PS', 'Both')) ──
        # Practice Alignment multiselect: dynamically loaded from active PS/Both users.
        # Brace notation stripped before display.
        try:
            _pa_rows = db.execute(text("""
                SELECT DISTINCT
                    TRIM(REPLACE(REPLACE(REPLACE(REPLACE(
                        COALESCE(practice_alignment, ''),
                        '{',''),'}',''),'"',''),chr(39),'')) AS pa
                FROM clockify_users
                WHERE status = 'active'
                  AND practice_area IN ('PS', 'Both')
                  AND daily_capacity > 0
                  AND TRIM(REPLACE(REPLACE(REPLACE(REPLACE(
                        COALESCE(practice_alignment, ''),
                        '{',''),'}',''),'"',''),chr(39),'')) != ''
                ORDER BY 1
            """)).fetchall()
            _pa_options = [r[0] for r in _pa_rows]
        except Exception:
            _pa_options = []

        _cap_practice = st.multiselect(
            "Practice Alignment",
            _pa_options,
            default=[],
            key="cap_practice",
        )

        # ── Build and execute the capacity query ──
        # Base: ALL active PS staff (practice_area IN ('PS', 'Both')).
        # LEFT JOIN to ps_resource_forecast_v2 for the 8-week forward window.
        # Users with no forecast rows appear once per week as Available / Unassigned / 0h.
        try:
            # Practice Alignment filter: when selections are present, add to the ps_users CTE.
            if _cap_practice:
                _practice_filter = (
                    "AND TRIM(REPLACE(REPLACE(REPLACE(REPLACE("
                    "COALESCE(practice_alignment,''),'{',''),'}',''),'\"',''),chr(39),''))"
                    " = ANY(:alignments)"
                )
                _cap_params = {"alignments": _cap_practice}
            else:
                _practice_filter = ""
                _cap_params = {}

            _cap_query = text(f"""
                WITH ps_users AS (
                    SELECT
                        clockify_user_id,
                        name AS user_name,
                        daily_capacity * 5 AS weekly_capacity,
                        TRIM(REPLACE(REPLACE(REPLACE(REPLACE(
                            COALESCE(practice_alignment, ''),
                            '{{',''),'}}',''),'"',''),chr(39),'')) AS practice_alignment
                    FROM clockify_users
                    WHERE status = 'active'
                      AND practice_area IN ('PS', 'Both')
                      AND daily_capacity > 0
                      AND NOT COALESCE(reporting_excluded, FALSE)
                      AND (time_submission IS NULL OR UPPER(TRIM(time_submission)) != 'NO')
                      {_practice_filter}
                ),
                week_spine AS (
                    SELECT generate_series(
                        DATE_TRUNC('week', CURRENT_DATE)::DATE,
                        DATE_TRUNC('week', CURRENT_DATE)::DATE + INTERVAL '7 weeks',
                        '1 week'::INTERVAL
                    )::DATE AS week_start
                ),
                base AS (
                    SELECT u.user_name, u.practice_alignment, u.weekly_capacity, w.week_start
                    FROM ps_users u CROSS JOIN week_spine w
                )
                SELECT
                    b.user_name          AS "Resource",
                    b.practice_alignment AS "Practice Alignment",
                    COALESCE(f.client_name,  'Available')   AS "Client",
                    COALESCE(f.project_name, 'Unassigned')  AS "Project",
                    b.week_start         AS "Week Start",
                    COALESCE(f.hours, 0)          AS "Hours",
                    COALESCE(f.allocation_pct, 0) AS "Allocation %",
                    COALESCE(f.is_actual, FALSE)  AS "Is Actual"
                FROM base b
                LEFT JOIN ps_resource_forecast_v2 f
                    ON  f.user_name  = b.user_name
                    AND f.week_start = b.week_start
                    AND f.hours > 0
                ORDER BY b.user_name, b.week_start, "Client"
            """)

            _cap_rows = db.execute(_cap_query, _cap_params).fetchall()

            if _cap_rows:
                _cap_df = pd.DataFrame(
                    _cap_rows,
                    columns=["Resource", "Practice Alignment", "Client", "Project",
                             "Week Start", "Hours", "Allocation %", "Is Actual"],
                )
                st.dataframe(_cap_df, use_container_width=True, hide_index=True)
                # Headcount: distinct PS staff currently shown
                _ps_headcount = _cap_df["Resource"].nunique()
                st.caption(f"Showing {_ps_headcount} PS staff member{'s' if _ps_headcount != 1 else ''}")
            else:
                st.info("No capacity data for the selected filters.")
        except Exception as _cap_e:
            st.warning(f"Could not load capacity data: {_cap_e}")

        st.divider()
        st.subheader("Add / Update Entry")
        _active_users = db.execute(text(
            "SELECT DISTINCT name FROM clockify_users WHERE status='active' ORDER BY name"
        )).fetchall()
        with st.form("capacity_entry_form"):
            _cc1, _cc2 = st.columns(2)
            with _cc1:
                _cap_user = st.selectbox("Resource", [r[0] for r in _active_users])
                _cap_client = st.text_input("Client Name")
                _cap_project = st.text_input("Project Name")
            with _cc2:
                _cap_week = st.date_input("Week Start", value=_cap_monday.date())
                _cap_hours = st.number_input("Hours", min_value=0.0, max_value=40.0, value=0.0, step=0.5)
                _cap_alloc = st.number_input("Allocation %", min_value=0, max_value=100, value=0)
            if st.form_submit_button("Save", type="primary"):
                try:
                    _cu = db.execute(text(
                        "SELECT clockify_user_id FROM clockify_users WHERE name=:n AND status='active' LIMIT 1"
                    ), {"n": _cap_user}).scalar()
                    db.execute(text("""
                        INSERT INTO ps_resource_forecast_v2
                            (clockify_user_id, user_name, client_name, project_name, week_start, hours, allocation_pct, is_actual)
                        VALUES (:uid, :uname, :client, :proj, :ws, :hrs, :alloc, FALSE)
                        ON CONFLICT (clockify_user_id, client_name, project_name, week_start)
                        DO UPDATE SET hours=EXCLUDED.hours, allocation_pct=EXCLUDED.allocation_pct
                    """), {"uid": _cu, "uname": _cap_user, "client": _cap_client, "proj": _cap_project,
                           "ws": _cap_week, "hrs": _cap_hours, "alloc": _cap_alloc})
                    db.commit()
                    st.success("Saved.")
                    st.rerun()
                except Exception as _ce:
                    db.rollback()
                    st.error(f"Save failed: {_ce}")

    with st.expander("⚙️ Forecast Model Config", expanded=False):
        try:
            _cfg_rows = db.execute(text("SELECT key, value FROM forecast_config")).fetchall()
            _cfg = {r[0]: float(r[1]) for r in _cfg_rows}
        except Exception:
            _cfg = {}

        w_hours = st.slider("Historical Hours Weight", 0.0, 1.0, step=0.05,
            value=_cfg.get('weight_historical_hours', 0.50), key="fc_w_hours")
        w_jira = st.slider("Jira Velocity Weight", 0.0, 1.0, step=0.05,
            value=_cfg.get('weight_jira_velocity', 0.30), key="fc_w_jira")
        w_pm = round(1.0 - w_hours - w_jira, 10)
        st.metric("PM Forecast Weight (auto)", f"{w_pm:.2f}")
        if abs(w_hours + w_jira + max(w_pm, 0.0) - 1.0) > 1e-9:
            st.warning("Weights do not sum to 1.0 — adjust sliders.")

        seasonal_on = st.toggle("Seasonal Correction",
            value=int(_cfg.get('seasonal_correction_enabled', 1)) == 1, key="fc_seasonal")
        lookback_default = st.number_input("Lookback Weeks (default)", min_value=1, max_value=52,
            value=int(_cfg.get('lookback_weeks_default', 8)), key="fc_lookback_def")
        lookback_min = st.number_input("Lookback Weeks (min data)", min_value=1, max_value=26,
            value=int(_cfg.get('lookback_weeks_min_data', 4)), key="fc_lookback_min")
        decay_weeks = st.number_input("Decay Start Weeks", min_value=0.5, max_value=8.0, step=0.5,
            value=float(_cfg.get('decay_start_weeks', 2.0)), key="fc_decay")

        if st.button("💾 Save", type="primary", key="fc_save"):
            updates = {
                'weight_historical_hours':    w_hours,
                'weight_jira_velocity':       w_jira,
                'weight_pm_forecast':         max(w_pm, 0.0),
                'seasonal_correction_enabled': 1 if seasonal_on else 0,
                'decay_start_weeks':          decay_weeks,
                'lookback_weeks_default':     lookback_default,
                'lookback_weeks_min_data':    lookback_min,
            }
            for k, v in updates.items():
                db.execute(text("UPDATE forecast_config SET value = :val, updated_at = NOW() WHERE key = :key"),
                    {"val": v, "key": k})
            db.commit()
            st.success("Forecast config saved.")
            st.rerun()

    st.divider()
    st.caption("Advanced Tools")
    adv_tab1, adv_tab2, adv_tab3, adv_tab4, adv_tab5 = st.tabs([
        "📊 Dashboard", "⚙️ Settings", "📐 Extensions", "▶️ Run Forecast", "⚖️ Model Config"
    ])
    with adv_tab1:
        from src.resource_forecast_tab6 import render as render_rf_tab6
        render_rf_tab6(engine)
    with adv_tab2:
        from src.resource_forecast_controls import _render_settings
        _render_settings(engine)
    with adv_tab3:
        from src.resource_forecast_controls import _render_extensions
        _render_extensions(engine)
    with adv_tab4:
        from src.resource_forecast_controls import _render_run_forecast
        _render_run_forecast(engine)
    with adv_tab5:
        from src.resource_forecast_controls import _render_model_config
        _render_model_config(db)

db.close()